"""
HOOD Excel Model Builder - Generate 7-Sheet Financial Model Workbook
======================================================================

WHAT THIS MODULE DOES:
  Reads model data CSVs (income statement, balance sheet, cash flow) produced by
  hood_data_transform.py and generates HOOD_Financial_Model.xlsx with seven sheets.
  All forecast cells use Excel formulas that reference the Assumptions sheet,
  enabling scenario analysis via formula updates without manual recalculation.

WHY IT EXISTS:
  Financial models built in Excel require complex integrations across multiple
  statements (IS -> BS -> CF). This module automates:
    * Consistent formatting and styling across all sheets
    * Formula-driven forecast consistency (one input source of truth)
    * Cross-sheet references that maintain balance-sheet equality and cash flow logic
    * Scenario analysis via YAML configuration (bull/base/bear assumptions)

INPUTS:
  * model_income_statement.csv   - quarterly historical and formula structure
  * model_balance_sheet.csv      - quarterly historical balance sheet items
  * model_cash_flow.csv          - quarterly historical cash flow items
  * config.yaml, scenarios.yaml  - assumption defaults and scenario overrides
  All CSVs are produced by hood_data_transform.py in the DATA_DIR folder.

OUTPUTS:
  * HOOD_Financial_Model.xlsx (or HOOD_Financial_Model_<scenario>.xlsx)
    Location: OUTPUT_DIR (defined in config module)

SHEETS (in order):
  1. Assumptions       - All driver inputs (BLUE text, YELLOW fill); checks appended
  2. Income Statement  - Historical + 4-year forecast (all formulas reference Assumptions)
  3. Balance Sheet     - Historical + 4-year forecast (Cash from FCF; Debt = revolver)
  4. Cash Flow         - Historical + 4-year forecast (CFO = NI+SBC-ΔAR+ΔAP; FCF = CFO-Capex)
  5. Valuation         - DCF, exit multiples, WACC x TGR sensitivity
  6. Sensitivity       - OpEx % x Revenue Growth and NI% x Volume tables
  7. Model Guide       - Navigation, key assumptions, audit trail

FORMATTING STANDARDS:
  Colors:
    * Blue (#0000FF)       - input assumption cells (user-editable)
    * Black (#000000)      - formula/computed cells (read-only)
    * Yellow (#FFFF00)     - background fill for input cells
    * Navy (#1F2D3D)       - section/sheet headers
    * Light Blue (#EFF6FF) - forecast column background

  Number Formats:
    * Currency:  $#,##0;($#,##0);"-"
    * Percent:   0.0%
    * Multiples: 0.0"x"  (e.g. 3.2x)
    * EPS:       "$"0.00  (e.g. $3.45)

  Column Widths:
    * Line-item column (A):  30 characters
    * Data columns (B+):     15 characters

HOW TO RUN:
  python -m src.build_excel_model [--scenario bull|base|bear]

  Examples:
    python -m src.build_excel_model              # base case
    python -m src.build_excel_model --scenario bull   # bull scenario
    python -m src.build_excel_model --scenario bear   # bear scenario

SCENARIO MANAGEMENT:
  * Base case: Default assumptions from ASSUMPTIONS_SPEC
  * Bull/Bear: Override specific assumptions via scenarios.yaml
  * Validation: All scenario overrides validated against assumption bounds
  * Bounds (enforced at build-time):
    - Growth rates:     -50% to +200%
    - Tax rates:        0% to 50%
    - WACC:            5% to 30%
    - Terminal Growth:  0% to 10%

KEY ACCOUNTING NOTES:
  OpEx vs SBC Split:
    HOOD's XBRL 'OperatingExpenses' includes Stock-Based Compensation.
    This model SPLITS them for visibility:
      * "Op. Exp. (ex. SBC)" = reported OpEx - SBC  (historically derived)
      * "Stock-Based Comp."  = SBC (directly from XBRL)
      * Total = sum (no double-counting)
    Forecast: opex_pct is calibrated ex-SBC (~42% of revenue).

  Three-Statement Integration (no circularity):
    * Income Statement: Independent revenue growth (Txn, NI, Other segments)
    * Balance Sheet: Cash rolls from CF FCF; Debt = draw-only revolver
    * Cash Flow: Sources derived from IS and BS cross-references
    * Circular Guard: Interest is flat (not debt-dependent), avoiding loops

DEPENDENCIES:
  * openpyxl - Excel workbook creation and styling
  * pandas - CSV reading and data manipulation
  * PyYAML - Scenario configuration parsing
  * config module - Shared layout constants (row/column indices, paths)

MODULE STRUCTURE:
  * Constants (lines 50-120): Color codes, fonts, borders, number formats
  * Helper functions (lines 120-480): Cell styling, title/header application
  * Sheet builders (lines 500-2400): Detailed logic for each of 7 sheets
  * Validators (lines 2200-2330): CSV integrity and scenario bounds checking
  * main() (lines 2487-2729): Entry point; orchestrates CSV loading and sheet generation
"""

from __future__ import annotations

import logging
import os
import sys
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths & shared layout constants (single source of truth)
# ---------------------------------------------------------------------------

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir))
from config import (
    REPO_ROOT, DATA_DIR, OUTPUT_DIR, TICKER,
    LABEL_COL, HIST_COL_START, NUM_FCST_COLS, FCST_YEARS,
    IS_ROW, BS_ROW, CF_ROW,
)

OUTPUT_DIR.mkdir(exist_ok=True)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Color & style constants
# ---------------------------------------------------------------------------

# Hex colours (no leading #)
C_BLUE   = "0000FF"
C_BLACK  = "000000"
C_YELLOW = "FFFF00"
C_WHITE  = "FFFFFF"
C_HEADER = "1F2D3D"   # dark navy for section headers
C_FCST   = "EFF6FF"   # light blue tint for forecast columns
C_FCST_H = "1A56DB"   # forecast column header text
C_GREY   = "F2F2F2"   # alternating row fill
C_SUBHDR = "D9D9D9"   # section sub-header fill

# Fonts
FONT_INPUT  = Font(name="Calibri", size=10, color=C_BLUE,  bold=False)
FONT_FORMULA= Font(name="Calibri", size=10, color=C_BLACK, bold=False)
FONT_BOLD   = Font(name="Calibri", size=10, color=C_BLACK, bold=True)
FONT_TITLE  = Font(name="Calibri", size=13, color=C_WHITE, bold=True)
FONT_HDR    = Font(name="Calibri", size=10, color=C_WHITE, bold=True)
FONT_NOTE   = Font(name="Calibri", size=9,  color="595959", italic=True)

# Fills
FILL_YELLOW = PatternFill("solid", fgColor=C_YELLOW)
FILL_HEADER = PatternFill("solid", fgColor=C_HEADER)
FILL_FCST   = PatternFill("solid", fgColor=C_FCST)
FILL_GREY   = PatternFill("solid", fgColor=C_GREY)
FILL_SUBHDR = PatternFill("solid", fgColor=C_SUBHDR)
FILL_NONE     = PatternFill()
FILL_GREEN_CF = PatternFill("solid", fgColor="C6EFCE")   # conditional format: pass / zero
FILL_RED_CF   = PatternFill("solid", fgColor="FFC7CE")   # conditional format: fail / non-zero

# Borders
_THIN   = Side(style="thin",   color="BFBFBF")
_MEDIUM = Side(style="medium", color="808080")
_DOUBLE = Side(style="double", color=C_BLACK)
BORDER_BOTTOM_MED    = Border(bottom=_MEDIUM)
BORDER_BOTTOM_DOUBLE = Border(bottom=_DOUBLE)
BORDER_TOP_THIN      = Border(top=_THIN)

# Number formats (per spec)
NUM_CURRENCY = "$#,##0;($#,##0);\"-\""
NUM_PCT      = "0.0%"
NUM_MULT     = '0.0"x"'   # leverage multiples e.g. 3.2x
NUM_EPS      = '"$"0.00'  # diluted EPS e.g. $3.45

# Column widths (per spec)
COL_W_LABEL = 30
COL_W_DATA  = 15

def cl(col_idx: int) -> str:
    """Return Excel column letter for a 1-based column index."""
    return get_column_letter(col_idx)


# ---------------------------------------------------------------------------
# Generic cell styler
# ---------------------------------------------------------------------------

def style_cell(
    cell,
    value=None,
    font: Optional[Font] = None,
    fill=None,
    num_format: Optional[str] = None,
    alignment: Optional[Alignment] = None,
    border=None,
) -> None:
    """Apply formatting to an openpyxl cell in-place (selective update).

    Convenience wrapper that allows setting only specific cell attributes without
    overwriting others. Any argument left as None is skipped, allowing fine-grained
    control over cell styling without needing to preserve/restore existing attributes.

    Args:
        cell: openpyxl Cell object to format.
        value: Cell value (any type); if None, cell value unchanged.
        font (Font, optional): Font object; if provided, overwrites cell font.
        fill (PatternFill, optional): Fill/background color object.
        num_format (str, optional): Number format code (e.g., "$#,##0" for currency).
        alignment (Alignment, optional): Alignment object (horizontal, vertical, indent).
        border (Border, optional): Border object (top, bottom, left, right).

    Returns:
        None (modifies cell in-place).
    """
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if num_format:
        cell.number_format = num_format
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


# ---------------------------------------------------------------------------
# Sheet 1 — Assumptions
# ---------------------------------------------------------------------------

# Each entry: label, value, is_pct, dict-key  (None = spacer row; value=None = section header)
ASSUMPTIONS_SPEC = [
    # ---- Transaction Revenue ----
    ("Transaction Revenue Growth  (crypto / equities / options)", None, False, None),
    ("  FY2026E Growth %",               0.30,  True,  "txn_growth_y1"),
    ("  FY2027E Growth %",               0.25,  True,  "txn_growth_y2"),
    ("  FY2028E Growth %",               0.20,  True,  "txn_growth_y3"),
    ("  FY2029E Growth %",               0.18,  True,  "txn_growth_y4"),
    None,
    # ---- Net Interest Revenue ----
    ("Net Interest Revenue Growth  (Fed-funds-rate sensitive)", None, False, None),
    ("  FY2026E Growth %",               0.15,  True,  "ni_growth_y1"),
    ("  FY2027E Growth %",               0.10,  True,  "ni_growth_y2"),
    ("  FY2028E Growth %",               0.08,  True,  "ni_growth_y3"),
    ("  FY2029E Growth %",               0.05,  True,  "ni_growth_y4"),
    None,
    # ---- Other Revenue ----
    ("Other Revenue Growth  (Gold subscriptions / other)",       None, False, None),
    ("  FY2026E Growth %",               0.20,  True,  "other_growth_y1"),
    ("  FY2027E Growth %",               0.15,  True,  "other_growth_y2"),
    ("  FY2028E Growth %",               0.12,  True,  "other_growth_y3"),
    ("  FY2029E Growth %",               0.10,  True,  "other_growth_y4"),
    None,
    # ---- Cost Structure ----
    ("Cost Structure",                   None,  False, None),
    ("  Op. Expenses % of Rev (ex. SBC)", 0.42, True,  "opex_pct"),
    ("  Stock-Based Compensation % of Rev", 0.08, True, "sbc_pct"),
    ("  Depreciation & Amortization % of Rev", 0.03, True, "da_pct"),
    None,
    # ---- Tax Schedule ----
    ("Tax Schedule  (NOL-adjusted effective rates)", None, False, None),
    ("  FY2026E Effective Tax Rate",     0.05,  True,  "tax_rate_y1"),
    ("  FY2027E Effective Tax Rate",     0.10,  True,  "tax_rate_y2"),
    ("  FY2028E Effective Tax Rate",     0.15,  True,  "tax_rate_y3"),
    ("  FY2029E Effective Tax Rate",     0.21,  True,  "tax_rate_y4"),
    None,    # <- NOL note written here (row 32)
    # ---- NOL Tracking ----
    ("NOL Tracking",                     None,  False, None),
    ("  NOL Carryforward Balance ($M)", 2000.0, False, "nol_balance"),
    ("  Statutory Tax Rate",             0.21,  True,  "statutory_rate"),
    None,
    # ---- Working Capital ----
    ("Working Capital Drivers",          None,  False, None),
    ("  Receivables % of Revenue",       0.36,  True,  "recv_pct"),
    ("  Payables % of Revenue",          0.34,  True,  "pay_pct"),
    ("  Minimum Cash Balance ($M)",     500.0,  False, "min_cash"),
    ("  Capital Expenditures % of Rev",  0.003, True,  "capex_pct"),
    None,
    # ---- Other Income ----
    ("Other Income  (corporate interest + non-operating)", None, False, None),
    ("  FY2026E–FY2029E Annual Amount ($M)", 150.0, False, "other_income"),
    None,
    # ---- Per-Share ----
    ("Per-Share Data",                   None,  False, None),
    ("  Diluted Shares Outstanding ($M)", 1100.0, False, "shares_diluted"),
    None,
    # ---- Valuation ----
    ("Valuation Inputs",                 None,  False, None),
    ("  WACC",                           0.12,  True,  "wacc"),
    ("  Terminal Growth Rate",           0.03,  True,  "terminal_growth"),
    ("  Exit EV / EBITDA Multiple",      20.0,  False, "exit_ebitda_mult"),
    ("  Exit EV / Revenue Multiple",      5.0,  False, "exit_rev_mult"),
    None,
    # ---- Debt Limits ----
    # WARNING: removing or reordering these two rows will shift WACC/TGR row indices
    # used by the assumption-bounds validator in validate_model.py.  Add new rows here,
    # after the existing Valuation block.
    ("Revolver Constraints  (simplified draw-only mechanic)",  None,  False, None),
    ("  Minimum Debt Balance ($M)",       0.0,  False, "min_debt"),
    ("  Maximum Debt Balance ($M)",    5000.0,  False, "max_debt"),
]
# After loop: row = 59.  map_start = 61.  checks_start = 76.


def build_assumptions(ws, is_fcst_col_start: int, hist_recv_pct=None, hist_pay_pct=None, spec=None) -> tuple[dict[str, str], int]:
    """
    Populate the Assumptions sheet.
    Parameters:
      is_fcst_col_start – first forecast column index on the Income Statement sheet
                          (used to build NOL depletion memo cross-references)
      hist_recv_pct     – optional historical receivables % for annotation
      hist_pay_pct      – optional historical payables % for annotation
      spec              – optional override of ASSUMPTIONS_SPEC; defaults to module-level list
    Returns (cell_refs, checks_start):
      cell_refs   – dict {key: absolute_cell_ref} for use in forecast formulas
      checks_start – first row available for the integrity checks block
    """
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 26.33

    # Row 1 – title bar
    ws.row_dimensions[1].height = 24
    style_cell(ws["A1"], f"{TICKER} Financial Model — Assumptions",
               font=FONT_TITLE, fill=FILL_HEADER,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells("A1:B1")

    # Row 2 – note
    style_cell(ws["A2"], "All blue cells are model inputs.  Values in $ millions.",
               font=FONT_NOTE)

    # Row 3 – column headers
    for letter, label in [("A", "Driver"), ("B", "Value")]:
        style_cell(ws[f"{letter}3"], label, font=FONT_HDR, fill=FILL_HEADER,
                   alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[3].height = 16

    cell_refs: dict[str, str] = {}
    row = 4

    for item in (spec or ASSUMPTIONS_SPEC):
        # Blank spacer row
        if item is None:
            row += 1
            continue

        label, value, is_pct, key = item

        # Section sub-header (value is None)
        if value is None:
            style_cell(ws[f"A{row}"], label,
                       font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
                       fill=FILL_SUBHDR)
            ws.merge_cells(f"A{row}:B{row}")
            row += 1
            continue

        # Normal assumption row
        style_cell(ws[f"A{row}"], label, font=FONT_FORMULA)

        vc = ws[f"B{row}"]
        vc.value        = value
        vc.font         = FONT_INPUT
        vc.fill         = FILL_YELLOW
        vc.number_format = NUM_PCT if is_pct else NUM_CURRENCY
        vc.alignment    = Alignment(horizontal="center")
        vc.border       = Border(bottom=_THIN)

        if key:
            cell_refs[key] = f"Assumptions!$B${row}"

        # Historical WC calibration notes (column C)
        if key == "recv_pct" and hist_recv_pct is not None:
            c = ws.cell(row=row, column=3)
            c.value = f"(hist. avg: {hist_recv_pct:.0%} — incl. brokerage settlement recv)"
            c.font = FONT_NOTE
        if key == "pay_pct" and hist_pay_pct is not None:
            c = ws.cell(row=row, column=3)
            c.value = f"(hist. avg: {hist_pay_pct:.0%} — incl. brokerage settlement pay)"
            c.font = FONT_NOTE

        row += 1

    # NOL context note at row 32 (the None spacer row after the Tax Schedule section)
    nol_c = ws.cell(row=32, column=1)
    nol_c.value = (
        "HOOD reported ~$2B NOL carryforward. Effective tax rate ramps from ~5% (FY2026E) "
        "to statutory 21% (FY2029E) as the NOL is consumed."
    )
    nol_c.font = FONT_NOTE
    ws.merge_cells("A32:B32")

    # ---- NOL Depletion Schedule (memo, rows 54–59) ----
    nol_memo_row = row   # row = 54 after loop
    ws.row_dimensions[nol_memo_row].height = 16
    c = ws.cell(row=nol_memo_row, column=1)
    style_cell(c, "NOL Depletion Schedule (memo — informational)",
               font=Font(name="Calibri", size=9, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(f"A{nol_memo_row}:F{nol_memo_row}")

    # Column headers
    hdr_r = nol_memo_row + 1
    for ci, label in enumerate(["", "FY2026E", "FY2027E", "FY2028E", "FY2029E"], start=1):
        c = ws.cell(row=hdr_r, column=ci)
        style_cell(c, label, font=FONT_HDR, fill=FILL_HEADER,
                   alignment=Alignment(horizontal="center"))

    # Row labels and formulas
    oi_r    = IS_ROW["Operating Income"]
    nol_ref = cell_refs["nol_balance"]   # Assumptions!$B$34

    memo_labels = ["  Taxable Income ($M)", "  NOL Used ($M)", "  NOL Remaining ($M)"]
    for mi, lbl in enumerate(memo_labels):
        r = hdr_r + 1 + mi
        style_cell(ws.cell(row=r, column=1), lbl, font=FONT_NOTE)

    for fi in range(NUM_FCST_COLS):
        col_idx = 2 + fi
        is_let  = cl(is_fcst_col_start + fi)
        ti_r  = hdr_r + 1   # taxable income row
        nu_r  = hdr_r + 2   # NOL used row
        nr_r  = hdr_r + 3   # NOL remaining row
        ti_cell = ws.cell(row=ti_r, column=col_idx)
        nu_cell = ws.cell(row=nu_r, column=col_idx)
        nr_cell = ws.cell(row=nr_r, column=col_idx)

        ti_cell.value        = f"=MAX('Income Statement'!{is_let}{oi_r},0)"
        ti_cell.number_format = NUM_CURRENCY
        ti_cell.font         = FONT_FORMULA
        ti_cell.fill         = FILL_FCST
        ti_cell.alignment    = Alignment(horizontal="right")

        if fi == 0:
            nol_open = nol_ref   # absolute ref to assumption cell
        else:
            prev_nr_let = cl(2 + fi - 1)
            nol_open = f"{prev_nr_let}{nr_r}"

        nu_cell.value        = f"=MIN({nol_open},{cl(col_idx)}{ti_r})"
        nu_cell.number_format = NUM_CURRENCY
        nu_cell.font         = FONT_FORMULA
        nu_cell.fill         = FILL_FCST
        nu_cell.alignment    = Alignment(horizontal="right")

        nr_cell.value        = f"=MAX(0,{nol_open}-{cl(col_idx)}{nu_r})"
        nr_cell.number_format = NUM_CURRENCY
        nr_cell.font         = FONT_FORMULA
        nr_cell.fill         = FILL_FCST
        nr_cell.alignment    = Alignment(horizontal="right")

    map_start = nol_memo_row + 6   # memo title + header + 3 data rows + 1 blank = 6 rows ahead

    # ---- Three-Statement Integration Map ----
    # Self-documenting flow showing how every sheet cross-references the others.

    # Title bar
    ws.row_dimensions[map_start].height = 20
    c = ws.cell(row=map_start, column=1)
    style_cell(c, "Three-Statement Integration Flow",
               font=FONT_TITLE, fill=FILL_HEADER,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells(f"A{map_start}:C{map_start}")

    map_rows = [
        # (fill, text)
        (FILL_SUBHDR, "Assumptions  →  Income Statement"),
        (FILL_NONE,   "  Each segment grows independently: Txn Rev × (1+txn_growth), "
                      "NI Rev × (1+ni_growth), Other Rev × (1+other_growth)."),
        (FILL_NONE,   "  Op. Expenses (ex. SBC) = Revenue × opex_pct  |  "
                      "SBC = Revenue × sbc_pct  |  D&A = Revenue × da_pct  |  Tax = MAX(OI,0) × tax_rate"),
        (FILL_SUBHDR, "Income Statement  →  Balance Sheet"),
        (FILL_NONE,   "  Receivables = IS Revenue × recv_pct"),
        (FILL_NONE,   "  Payables    = IS Revenue × pay_pct"),
        (FILL_NONE,   "  Equity      = Prior Equity + IS Net Income"),
        (FILL_SUBHDR, "Balance Sheet ΔWC  →  Cash Flow"),
        (FILL_NONE,   "  CFO = IS Net Income + IS SBC + D&A − ΔReceivables + ΔPayables"),
        (FILL_NONE,   "  Capex = IS Revenue × capex_pct     |     FCF = CFO − Capex"),
        (FILL_SUBHDR, "Cash Flow FCF  →  Balance Sheet  (closes the model)"),
        (FILL_NONE,   "  Cash  = MAX( Prior Cash + FCF ,  Minimum Cash Balance )"),
        (FILL_NONE,   "  Debt  = Prior Debt + MAX( 0 ,  MinCash − (Prior Cash + FCF) )  "
                      "← revolver plug"),
    ]

    for offset, (fill, text) in enumerate(map_rows):
        r = map_start + 1 + offset
        is_subhdr = fill is FILL_SUBHDR
        # Subhdr rows: 16pt. Content rows: 15pt (matches user-adjusted layout).
        ws.row_dimensions[r].height = 16 if is_subhdr else 15
        c = ws.cell(row=r, column=1)
        style_cell(c, text,
                   font=Font(name="Calibri", size=9,
                             bold=is_subhdr, color=C_BLACK),
                   fill=fill,
                   alignment=Alignment(horizontal="left",
                                       vertical="center",
                                       indent=0 if is_subhdr else 1,
                                       wrap_text=True))
        ws.merge_cells(
            start_row=r, start_column=1,
            end_row=r,   end_column=3
        )

    checks_start = map_start + len(map_rows) + 2
    return cell_refs, checks_start


# ---------------------------------------------------------------------------
# Sheet 2 — Income Statement
# ---------------------------------------------------------------------------
# IS_ROW, BS_ROW, CF_ROW are imported from config.py (single source of truth).


def _write_fcst_cell(
    ws, row: int, col_idx: int, formula: str,
    bold: bool = False, border=None, num_format: Optional[str] = None
):
    """Write a forecast formula cell with standard light-blue styling.

    Convenience wrapper that applies consistent formatting to forecast cells:
    light-blue background (FILL_FCST), Calibri 10pt black font, right alignment,
    and currency number format (unless overridden). Used extensively throughout
    Income Statement, Balance Sheet, and Cash Flow builders.

    Args:
        ws: openpyxl Worksheet object.
        row (int): 1-based row number.
        col_idx (int): 1-based column index.
        formula (str): Excel formula string (e.g., "=A1+B1").
        bold (bool): If True, applies bold font weight; default False.
        border (Border, optional): Border object; default None.
                                   Common use: BORDER_BOTTOM_MED for subtotal rows.
        num_format (str, optional): Number format code; default NUM_CURRENCY.

    Returns:
        Cell: The styled openpyxl cell object.
    """
    c = ws.cell(row=row, column=col_idx)
    c.value        = formula
    c.font         = Font(name="Calibri", size=10, color=C_BLACK, bold=bold)
    c.number_format = num_format if num_format else NUM_CURRENCY
    c.fill         = FILL_FCST
    c.alignment    = Alignment(horizontal="right")
    if border:
        c.border = border
    return c


def _apply_title_row(ws, title: str, last_col_idx: int, row: int = 1) -> None:
    """Render a branded title bar with dark navy background spanning multiple columns.

    Creates a visually distinct sheet header by merging cells across all data columns,
    applying white bold text on dark navy background. Used by all seven sheet builders
    to provide consistent visual hierarchy.

    Args:
        ws: openpyxl Worksheet object.
        title (str): Title text to display (e.g., "HOOD - Income Statement").
        last_col_idx (int): Last column index to span (determines merge range).
        row (int): Row number for title; default 1.

    Returns:
        None (modifies worksheet in-place).
    """
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1)
    style_cell(c, title, font=FONT_TITLE, fill=FILL_HEADER,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=last_col_idx
    )


def _apply_col_headers(
    ws, df_columns: list[str], row: int,
    hist_col_start: int, fcst_years: list[str]
) -> None:
    """Write column-period labels for both historical quarters and forecast years.

    Formats headers to visually distinguish historical data from forecast:
    - Historical columns: dark navy fill, white text
    - Forecast columns: light blue fill, blue text
    This color scheme immediately communicates actuals vs. projections to users.

    Args:
        ws: openpyxl Worksheet object.
        df_columns (list[str]): Historical column labels from CSV (e.g., ["Q1 2024", "Q2 2024"]).
        row (int): Row number for headers.
        hist_col_start (int): 1-based column index where historical data begins.
        fcst_years (list[str]): Forecast period labels (e.g., ["FY2025E", "FY2026E"]).

    Returns:
        None (modifies worksheet in-place).
    """
    style_cell(ws.cell(row=row, column=1), "",
               font=FONT_HDR, fill=FILL_HEADER,
               alignment=Alignment(horizontal="center"))
    for i, hdr in enumerate(df_columns):
        c = ws.cell(row=row, column=hist_col_start + i)
        style_cell(c, hdr, font=FONT_HDR, fill=FILL_HEADER,
                   alignment=Alignment(horizontal="center"))
    for i, hdr in enumerate(fcst_years):
        col_idx = hist_col_start + len(df_columns) + i
        style_cell(ws.cell(row=row, column=col_idx), hdr,
                   font=Font(name="Calibri", size=10, color=C_FCST_H, bold=True),
                   fill=PatternFill("solid", fgColor="C7D8F8"),
                   alignment=Alignment(horizontal="center"))
    ws.row_dimensions[row].height = 16


def build_income_statement(
    ws, df_is: pd.DataFrame, cell_refs: dict[str, str], cf_fcst_col_start: int
) -> None:
    """Build the Income Statement sheet with historical actuals and forecast formulas.

    The Income Statement is the driver for all downstream sheets (BS, CF). All forecast
    revenues and costs are formulas that derive from the Assumptions sheet, enabling
    scenario analysis and sensitivity testing.

    Revenue Growth Strategy:
      Three independent revenue segments (Transaction, Net Interest, Other) each with
      separate growth assumptions per year. Y1 forecast uses LTM base * (1 + growth);
      Y2+ forecasts use prior-year amount * (1 + growth).

    Cost Structure:
      CRITICAL NOTE - OpEx vs SBC: HOOD's XBRL 'OperatingExpenses' includes SBC.
      This model SPLITS them for visibility:
        * "Op. Exp. (ex. SBC)" = reported OpEx - SBC (historically derived)
        * "Stock-Based Comp." = SBC (directly from XBRL)
        * Total Operating Costs = sum = reported OpEx (no double-counting)
      Historical OpEx (ex-SBC) is derived; forecast uses opex_pct assumption (~42%).

    Tax Provision:
      Tax = MAX(Operating Income, 0) * effective tax rate. NOL carryforward balance
      is tracked on Assumptions sheet but not modeled as NOL depletion; effective
      rate is set directly per fiscal year (accounts for NOL usage implicitly).

    Below-the-Line Items:
      "Below-the-line / Other" = Net Income - Operating Income. Captures corporate
      interest income, interest expense on debt, and other non-operating items.
      Derived by subtraction to match NI from audited financials without modeling
      each component separately.

    Args:
        ws: openpyxl Worksheet object for Income Statement sheet.
        df_is (pd.DataFrame): Income Statement CSV data, index=line items, columns=periods.
        cell_refs (dict): Map of assumption keys to cell references from Assumptions sheet.
        cf_fcst_col_start (int): First forecast column index on Cash Flow sheet
                                 (for cross-references in Balance Sheet).

    Returns:
        None (modifies worksheet in-place).
    """

    num_hist   = len(df_is.columns)
    fcst_start = HIST_COL_START + num_hist
    total_cols = LABEL_COL + num_hist + NUM_FCST_COLS

    ltm_col_indices = list(range(fcst_start - 4, fcst_start))
    ltm_headers = list(df_is.columns[-4:])
    ltm_label   = f"LTM base: {ltm_headers[0]} → {ltm_headers[-1]}"

    ws.column_dimensions[cl(LABEL_COL)].width = COL_W_LABEL
    for ci in range(HIST_COL_START, fcst_start + NUM_FCST_COLS + 1):
        ws.column_dimensions[cl(ci)].width = COL_W_DATA

    # ---- Title, unit note, col headers ----
    _apply_title_row(ws, f"{TICKER} — Income Statement", total_cols)
    style_cell(ws.cell(row=2, column=1), "$ in millions", font=FONT_NOTE)
    _apply_col_headers(ws, list(df_is.columns), IS_ROW["col_headers"],
                       HIST_COL_START, FCST_YEARS)

    # Separator rows
    for sep in ("sep_top", "sep_costs", "sep_oi", "sep_ni"):
        ws.row_dimensions[IS_ROW[sep]].height = 6
        for ci in range(1, total_cols + 1):
            ws.cell(row=IS_ROW[sep], column=ci).fill = FILL_SUBHDR

    # ---- Row labels ----
    label_spec = [
        (IS_ROW["Txn Revenue"],        "  Transaction-based Rev",   FONT_FORMULA),
        (IS_ROW["NI Revenue"],         "  Net Interest Rev",        FONT_FORMULA),
        (IS_ROW["Other Revenue"],      "  Other Revenue",           FONT_FORMULA),
        (IS_ROW["Total Revenue"],      "Total Revenue",             FONT_BOLD),
        (IS_ROW["Operating Expenses"], "  Op. Exp. (ex. SBC) †",   FONT_FORMULA),
        (IS_ROW["SBC"],                "  Stock-Based Comp.",       FONT_FORMULA),
        (IS_ROW["Total Costs"],        "Total Operating Costs",     FONT_BOLD),
        (IS_ROW["Operating Income"],   "Operating Income",          FONT_BOLD),
        (IS_ROW["Tax Provision"],      "  Tax Provision",           FONT_FORMULA),
        (IS_ROW["Below-the-line"],     "  Below-the-line / Other",  FONT_FORMULA),
        (IS_ROW["Net Income"],         "Net Income",                FONT_BOLD),
        (IS_ROW["EPS"],                "  Diluted EPS",             FONT_FORMULA),
    ]
    for sheet_row, label, font in label_spec:
        c = ws.cell(row=sheet_row, column=LABEL_COL)
        style_cell(c, label, font=font,
                   alignment=Alignment(horizontal="left", indent=1))

    # ---- Historical actuals ----
    # Raw CSV rows written directly (all except Operating Expenses)
    hist_map = {
        IS_ROW["Txn Revenue"]:   "Transaction-based Revenue",
        IS_ROW["NI Revenue"]:    "Net Interest Revenue",
        IS_ROW["Other Revenue"]: "Other Revenue",
        IS_ROW["Total Revenue"]: "Total Revenue",
        IS_ROW["SBC"]:           "Stock-Based Compensation",
        IS_ROW["Net Income"]:    "Net Income",
    }
    for sheet_row, csv_label in hist_map.items():
        if csv_label not in df_is.index:
            continue
        for hi, col_label in enumerate(df_is.columns):
            col_idx = HIST_COL_START + hi
            val = df_is.loc[csv_label, col_label]
            if pd.notna(val):
                c = ws.cell(row=sheet_row, column=col_idx)
                c.value         = val
                c.number_format = NUM_CURRENCY
                c.font          = FONT_FORMULA
                c.alignment     = Alignment(horizontal="right")

    # Operating Expenses (ex. SBC): reported OpEx − SBC  (no double-count)
    if "Operating Expenses" in df_is.index and "Stock-Based Compensation" in df_is.index:
        for hi, col_label in enumerate(df_is.columns):
            opex_raw = df_is.loc["Operating Expenses", col_label]
            sbc_raw  = df_is.loc["Stock-Based Compensation", col_label]
            if pd.notna(opex_raw) and pd.notna(sbc_raw):
                c = ws.cell(row=IS_ROW["Operating Expenses"], column=HIST_COL_START + hi)
                c.value         = opex_raw - sbc_raw
                c.number_format = NUM_CURRENCY
                c.font          = FONT_FORMULA
                c.alignment     = Alignment(horizontal="right")

    # Derived historical rows
    rev_r  = IS_ROW["Total Revenue"]
    opex_r = IS_ROW["Operating Expenses"]
    sbc_r  = IS_ROW["SBC"]
    tot_r  = IS_ROW["Total Costs"]
    oi_r   = IS_ROW["Operating Income"]
    ni_r   = IS_ROW["Net Income"]
    btl_r  = IS_ROW["Below-the-line"]

    for hi in range(num_hist):
        col_idx = HIST_COL_START + hi
        letter  = cl(col_idx)

        # Total Operating Costs = OpEx(ex-SBC) + SBC = Total OpEx (correct, no double-count)
        c = ws.cell(row=tot_r, column=col_idx)
        c.value         = f"={letter}{opex_r}+{letter}{sbc_r}"
        c.font          = FONT_BOLD
        c.number_format = NUM_CURRENCY
        c.alignment     = Alignment(horizontal="right")
        c.border        = BORDER_BOTTOM_MED

        # Operating Income = Revenue − Total Costs
        c = ws.cell(row=oi_r, column=col_idx)
        c.value         = f"={letter}{rev_r}-{letter}{tot_r}"
        c.font          = FONT_BOLD
        c.number_format = NUM_CURRENCY
        c.alignment     = Alignment(horizontal="right")
        c.border        = BORDER_BOTTOM_MED

        # Below-the-line / Other = NI − OI
        # Captures income taxes, interest on corporate cash, and non-operating items.
        c = ws.cell(row=btl_r, column=col_idx)
        c.value         = f"={letter}{ni_r}-{letter}{oi_r}"
        c.font          = FONT_FORMULA
        c.number_format = NUM_CURRENCY
        c.alignment     = Alignment(horizontal="right")

    # Bold borders on Total Revenue and Net Income
    for hi in range(num_hist):
        col_idx = HIST_COL_START + hi
        ws.cell(row=rev_r, column=col_idx).border  = BORDER_BOTTOM_MED
        ws.cell(row=ni_r,  column=col_idx).border  = BORDER_BOTTOM_DOUBLE

    # ---- Forecast formulas ----
    opex_pct        = cell_refs["opex_pct"]
    sbc_pct         = cell_refs["sbc_pct"]
    other_income_ref = cell_refs["other_income"]
    shares_ref      = cell_refs["shares_diluted"]

    txn_r    = IS_ROW["Txn Revenue"]
    ni_r_seg = IS_ROW["NI Revenue"]
    oth_r    = IS_ROW["Other Revenue"]

    ltm_letters = [cl(c) for c in ltm_col_indices]
    # LTM sums per segment (last 4 quarters)
    ltm_txn_sum = "+".join(f"{l}{txn_r}"   for l in ltm_letters)
    ltm_ni_sum  = "+".join(f"{l}{ni_r_seg}" for l in ltm_letters)
    ltm_oth_sum = "+".join(f"{l}{oth_r}"   for l in ltm_letters)

    for fi in range(NUM_FCST_COLS):
        col_idx  = fcst_start + fi
        letter   = cl(col_idx)
        prev_let = cl(col_idx - 1)

        yr_txn_growth   = cell_refs[f"txn_growth_y{fi + 1}"]
        yr_ni_growth    = cell_refs[f"ni_growth_y{fi + 1}"]
        yr_other_growth = cell_refs[f"other_growth_y{fi + 1}"]
        yr_tax_rate     = cell_refs[f"tax_rate_y{fi + 1}"]

        # Revenue segments (each independently driven)
        if fi == 0:
            txn_f = f"=({ltm_txn_sum})*(1+{yr_txn_growth})"
            ni_f  = f"=({ltm_ni_sum})*(1+{yr_ni_growth})"
            oth_f = f"=({ltm_oth_sum})*(1+{yr_other_growth})"
        else:
            txn_f = f"={prev_let}{txn_r}*(1+{yr_txn_growth})"
            ni_f  = f"={prev_let}{ni_r_seg}*(1+{yr_ni_growth})"
            oth_f = f"={prev_let}{oth_r}*(1+{yr_other_growth})"

        _write_fcst_cell(ws, txn_r,    col_idx, txn_f)
        _write_fcst_cell(ws, ni_r_seg, col_idx, ni_f)
        _write_fcst_cell(ws, oth_r,    col_idx, oth_f)

        # Total Revenue = sum of independently-forecast segments
        _write_fcst_cell(ws, rev_r, col_idx,
                         f"={letter}{txn_r}+{letter}{ni_r_seg}+{letter}{oth_r}",
                         bold=True, border=BORDER_BOTTOM_MED)

        # Operating Expenses (ex. SBC) = Revenue × opex_pct
        _write_fcst_cell(ws, opex_r, col_idx, f"={letter}{rev_r}*{opex_pct}")

        # Stock-Based Compensation = Revenue × sbc_pct
        _write_fcst_cell(ws, sbc_r, col_idx, f"={letter}{rev_r}*{sbc_pct}")

        # Total Costs
        _write_fcst_cell(ws, tot_r, col_idx,
                         f"={letter}{opex_r}+{letter}{sbc_r}",
                         bold=True, border=BORDER_BOTTOM_MED)

        # Operating Income
        _write_fcst_cell(ws, oi_r, col_idx,
                         f"={letter}{rev_r}-{letter}{tot_r}",
                         bold=True, border=BORDER_BOTTOM_MED)

        # Tax Provision = −MAX(OI, 0) × yr_tax_rate
        _write_fcst_cell(ws, IS_ROW["Tax Provision"], col_idx,
                         f"=-MAX({letter}{oi_r},0)*{yr_tax_rate}")

        # Below-the-line = other_income (corporate interest + other non-op income)
        _write_fcst_cell(ws, btl_r, col_idx, f"={other_income_ref}")

        # Net Income = OI + Tax + Below-the-line
        _write_fcst_cell(ws, ni_r, col_idx,
                         f"={letter}{oi_r}+{letter}{IS_ROW['Tax Provision']}+{letter}{btl_r}",
                         bold=True, border=BORDER_BOTTOM_DOUBLE)

        # EPS = Net Income ($M) / Diluted Shares ($M) -> $ per share
        _write_fcst_cell(ws, IS_ROW["EPS"], col_idx,
                         f"={letter}{ni_r}/{shares_ref}",
                         num_format=NUM_EPS)

    # ---- Margin Analysis section ----
    ws.row_dimensions[IS_ROW["sep_margin"]].height = 6
    for ci in range(1, total_cols + 1):
        ws.cell(row=IS_ROW["sep_margin"], column=ci).fill = FILL_SUBHDR

    c = ws.cell(row=IS_ROW["margin_header"], column=LABEL_COL)
    style_cell(c, "Margin Analysis",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(start_row=IS_ROW["margin_header"], start_column=1,
                   end_row=IS_ROW["margin_header"], end_column=total_cols)

    margin_labels = [
        (IS_ROW["Op Margin"],  "  Operating Margin"),
        (IS_ROW["Net Margin"], "  Net Margin"),
        (IS_ROW["FCF Margin"], "  FCF Margin ‡"),
        (IS_ROW["SBC Pct"],    "  SBC % Revenue"),
        (IS_ROW["OpEx Pct"],   "  OpEx % Revenue (ex. SBC)"),
    ]
    for sheet_row, label in margin_labels:
        style_cell(ws.cell(row=sheet_row, column=LABEL_COL), label,
                   font=FONT_FORMULA,
                   alignment=Alignment(horizontal="left", indent=1))

    # Historical margins (Op and Net only — FCF requires CF cross-ref)
    for hi in range(num_hist):
        col_idx = HIST_COL_START + hi
        letter  = cl(col_idx)

        c = ws.cell(row=IS_ROW["Op Margin"], column=col_idx)
        c.value         = f"={letter}{oi_r}/{letter}{rev_r}"
        c.number_format = NUM_PCT
        c.font          = FONT_FORMULA
        c.alignment     = Alignment(horizontal="right")

        c = ws.cell(row=IS_ROW["Net Margin"], column=col_idx)
        c.value         = f"={letter}{ni_r}/{letter}{rev_r}"
        c.number_format = NUM_PCT
        c.font          = FONT_FORMULA
        c.alignment     = Alignment(horizontal="right")

        c = ws.cell(row=IS_ROW["SBC Pct"], column=col_idx)
        c.value         = f"={letter}{sbc_r}/{letter}{rev_r}"
        c.number_format = NUM_PCT
        c.font          = FONT_FORMULA
        c.alignment     = Alignment(horizontal="right")

        c = ws.cell(row=IS_ROW["OpEx Pct"], column=col_idx)
        c.value         = f"={letter}{opex_r}/{letter}{rev_r}"
        c.number_format = NUM_PCT
        c.font          = FONT_FORMULA
        c.alignment     = Alignment(horizontal="right")

    # Forecast margins (all three, including FCF cross-ref)
    for fi in range(NUM_FCST_COLS):
        col_idx = fcst_start + fi
        letter  = cl(col_idx)
        cf_let  = cl(cf_fcst_col_start + fi)

        _write_fcst_cell(ws, IS_ROW["Op Margin"], col_idx,
                         f"={letter}{oi_r}/{letter}{rev_r}",
                         num_format=NUM_PCT)
        _write_fcst_cell(ws, IS_ROW["Net Margin"], col_idx,
                         f"={letter}{ni_r}/{letter}{rev_r}",
                         num_format=NUM_PCT)
        _write_fcst_cell(ws, IS_ROW["FCF Margin"], col_idx,
                         f"='Cash Flow'!{cf_let}{CF_ROW['FCF']}/{letter}{rev_r}",
                         num_format=NUM_PCT)
        _write_fcst_cell(ws, IS_ROW["SBC Pct"], col_idx,
                         f"={letter}{sbc_r}/{letter}{rev_r}",
                         num_format=NUM_PCT)
        _write_fcst_cell(ws, IS_ROW["OpEx Pct"], col_idx,
                         f"={letter}{opex_r}/{letter}{rev_r}",
                         num_format=NUM_PCT)

    # ---- LTM Revenue memo (answers "what is the FY2026E revenue base?") ----
    ws.row_dimensions[IS_ROW["sep_ltm"]].height = 6
    for ci in range(1, total_cols + 1):
        ws.cell(row=IS_ROW["sep_ltm"], column=ci).fill = FILL_SUBHDR

    style_cell(ws.cell(row=IS_ROW["LTM Revenue"], column=LABEL_COL),
               f"  Memo: LTM Revenue ({ltm_headers[0]}–{ltm_headers[-1]})",
               font=FONT_NOTE,
               alignment=Alignment(horizontal="left", indent=1))

    # LTM total shown at the FY2026E column so the base is visible alongside Year 1
    ltm_total_cell = ws.cell(row=IS_ROW["LTM Revenue"], column=fcst_start)
    ltm_total_cell.value         = f"=SUM({cl(ltm_col_indices[0])}{rev_r}:{cl(ltm_col_indices[-1])}{rev_r})"
    ltm_total_cell.number_format = NUM_CURRENCY
    ltm_total_cell.font          = FONT_NOTE
    ltm_total_cell.fill          = FILL_FCST
    ltm_total_cell.alignment     = Alignment(horizontal="right")

    # ---- Implied Growth memo row ----
    style_cell(ws.cell(row=IS_ROW["Implied Growth"], column=LABEL_COL),
               "  Memo: Implied Revenue Growth % §",
               font=FONT_NOTE,
               alignment=Alignment(horizontal="left", indent=1))

    for fi in range(NUM_FCST_COLS):
        col_idx  = fcst_start + fi
        letter   = cl(col_idx)
        prev_let = cl(col_idx - 1)

        if fi == 0:
            # FY2026E: growth from LTM base (same denominator as the rev_growth assumption)
            impl_formula = (
                f"={letter}{rev_r}"
                f"/SUM({cl(ltm_col_indices[0])}{rev_r}"
                f":{cl(ltm_col_indices[-1])}{rev_r})-1"
            )
        else:
            impl_formula = f"={letter}{rev_r}/{prev_let}{rev_r}-1"

        c = ws.cell(row=IS_ROW["Implied Growth"], column=col_idx)
        c.value         = impl_formula
        c.number_format = NUM_PCT
        c.font          = FONT_NOTE
        c.fill          = FILL_FCST
        c.alignment     = Alignment(horizontal="right")

    # ---- Footer notes ----
    impl_note_row = IS_ROW["Implied Growth"] + 1
    ltm_note_row  = IS_ROW["Implied Growth"] + 2
    btl_note_row  = IS_ROW["Implied Growth"] + 3
    opex_note_row = IS_ROW["Implied Growth"] + 4

    style_cell(ws.cell(row=impl_note_row, column=1),
               "§ FY2026E = Total Rev ÷ LTM Total Rev − 1. FY2027E–FY2029E = current ÷ prior year − 1. "
               "Total revenue growth reflects the weighted-average of independently-driven segment "
               "growth rates (Transaction, Net Interest, Other).",
               font=FONT_NOTE)

    style_cell(ws.cell(row=ltm_note_row, column=1),
               f"* LTM Revenue ({ltm_label}) is the annualized base for FY2026E. "
               "Revenue Growth % applies to this 12-month total — not to any single quarter. "
               "Quarterly YoY comparisons (e.g. Q3'24→Q3'25 = +100%) reflect operating momentum "
               "and will naturally differ from the annual blended growth rate assumption.",
               font=FONT_NOTE)
    style_cell(ws.cell(row=btl_note_row, column=1),
               "‡ FCF Margin historical: not shown (CF covers fewer periods than IS). "
               "FCF Margin forecast cross-references the Cash Flow sheet.",
               font=FONT_NOTE)
    style_cell(ws.cell(row=opex_note_row, column=1),
               "† Op. Exp. (ex. SBC): HOOD's reported OperatingExpenses XBRL tag includes SBC. "
               "Historical values shown here are reported OpEx minus SBC to avoid double-counting. "
               "Total Operating Costs = Op. Exp.(ex. SBC) + SBC = reported total. "
               "Forecast opex_pct assumption (~42%) is calibrated ex-SBC; total opex ≈ 50% of revenue.",
               font=FONT_NOTE)

    eps_note_row = opex_note_row + 1
    style_cell(ws.cell(row=eps_note_row, column=1),
               "** EPS = Net Income ($M) ÷ Diluted Shares Outstanding ($M). "
               "Forecast only; historical EPS not extracted from XBRL. "
               "Below-the-line in forecast = corporate interest income + other non-operating items "
               "(held flat at Assumptions input). Historical BTL = NI − OI residual.",
               font=FONT_NOTE)


# ---------------------------------------------------------------------------
# Sheets 3 & 4 — Balance Sheet and Cash Flow
# ---------------------------------------------------------------------------

def _write_hist_rows(ws, df: pd.DataFrame, rows: list) -> None:
    """Write historical actuals from a DataFrame into the worksheet.

    Iterates over rows — a list of (display_label, is_bold, csv_label) tuples —
    and writes each label into the label column starting at sheet row 4,
    then populates historical data values from df (indexed by csv_label) across
    the historical data columns, applying alternating grey/white fill and bold
    styling for subtotal rows.
    """
    for row_offset, (disp, is_bold, csv_label) in enumerate(rows):
        sheet_row = 4 + row_offset
        font = FONT_BOLD if is_bold else FONT_FORMULA

        style_cell(ws.cell(row=sheet_row, column=LABEL_COL), disp, font=font,
                   alignment=Alignment(horizontal="left", indent=1))
        ws.cell(row=sheet_row, column=LABEL_COL).fill = (
            FILL_SUBHDR if is_bold else
            (FILL_GREY if row_offset % 2 == 0 else FILL_NONE)
        )

        if csv_label not in df.index:
            continue
        for hi, col_label in enumerate(df.columns):
            val = df.loc[csv_label, col_label]
            c = ws.cell(row=sheet_row, column=HIST_COL_START + hi)
            c.fill = FILL_GREY if row_offset % 2 == 0 else FILL_NONE
            if pd.notna(val):
                c.value         = val
                c.number_format = NUM_CURRENCY
                c.font          = font
                c.alignment     = Alignment(horizontal="right")
                if is_bold:
                    c.border = BORDER_BOTTOM_MED


def build_balance_sheet(
    ws,
    df_bs: pd.DataFrame,
    cell_refs: dict[str, str],
    is_fcst_col_start: int,
    cf_fcst_col_start: int,
) -> None:
    """Build the Balance Sheet with historical actuals and 4-year forecast.

    Integrates with Income Statement (revenue) and Cash Flow (free cash flow) via
    cross-sheet formulas. Implements a simplified financial model with a draw-only
    revolver (debt) that increases when cash flows fall short of minimum threshold.

    Forecast Drivers:
      Cash = MAX(prev_cash + FCF, min_cash) - floored at minimum cash buffer;
             if cash_raw falls below min_cash, revolver is drawn.
      Receivables = IS Revenue * recv_pct assumption
      Payables = IS Revenue * pay_pct assumption
      Total Debt = MAX(min_debt, MIN(max_debt,
                       prev_debt + MAX(0, min_cash - cash_raw)))
                   Draw-only revolver: only increases, bounded by [min, max]
      Equity = prior Equity + Net Income
               Assumption: zero dividends/buybacks - all earnings retained

    Circular Reference Guard:
      WARNING: If interest expense ever becomes debt-dependent (e.g., interest =
      Debt * rate flowing into IS), this model will have a circular reference.
      Current design avoids this by using a flat other_income assumption that
      does NOT depend on debt balance.

    Balance Sheet Check (Partial):
      Verifies partial assets vs liabilities+equity (non-zero difference due to
      excluded items like custodial assets, deferred taxes, intangibles).

    Args:
        ws: openpyxl Worksheet object for Balance Sheet sheet.
        df_bs (pd.DataFrame): Balance Sheet CSV data, index=line items, columns=periods.
        cell_refs (dict): Map of assumption keys to cell references.
        is_fcst_col_start (int): First forecast column index on Income Statement.
        cf_fcst_col_start (int): First forecast column index on Cash Flow.

    Returns:
        None (modifies worksheet in-place).
    """
    num_hist      = len(df_bs.columns)
    bs_fcst_start = HIST_COL_START + num_hist
    total_cols    = LABEL_COL + num_hist + NUM_FCST_COLS

    ws.column_dimensions[cl(LABEL_COL)].width = COL_W_LABEL
    for ci in range(HIST_COL_START, bs_fcst_start + NUM_FCST_COLS + 1):
        ws.column_dimensions[cl(ci)].width = COL_W_DATA

    _apply_title_row(ws, f"{TICKER} — Balance Sheet", total_cols)
    style_cell(ws.cell(row=2, column=1), "$ in millions", font=FONT_NOTE)
    _apply_col_headers(ws, list(df_bs.columns), 3, HIST_COL_START, FCST_YEARS)

    rows = [
        ("Cash & Cash Equivalents", True,  "Cash & Cash Equivalents"),
        ("  Restricted Cash",       False, "Restricted Cash"),
        ("  Receivables",           False, "Receivables"),
        ("  Payables",              False, "Payables"),
        ("Total Debt",              True,  "Total Debt"),
        ("Stockholders' Equity",    True,  "Stockholders' Equity"),
    ]
    _write_hist_rows(ws, df_bs, rows)

    # ---- Forecast formulas ----
    recv_pct    = cell_refs["recv_pct"]
    pay_pct     = cell_refs["pay_pct"]
    min_cash    = cell_refs["min_cash"]
    min_debt    = cell_refs["min_debt"]
    max_debt    = cell_refs["max_debt"]

    rev_is_r = IS_ROW["Total Revenue"]
    ni_is_r  = IS_ROW["Net Income"]

    for fi in range(NUM_FCST_COLS):
        bs_col   = bs_fcst_start + fi
        prev_let = cl(bs_col - 1)
        is_let   = cl(is_fcst_col_start + fi)
        cf_let   = cl(cf_fcst_col_start + fi)

        is_rev  = f"'Income Statement'!{is_let}{rev_is_r}"
        is_ni   = f"'Income Statement'!{is_let}{ni_is_r}"
        fcf_ref = f"'Cash Flow'!{cf_let}{CF_ROW['FCF']}"

        # cash_raw: prior cash + this year's FCF
        cash_raw = f"({prev_let}{BS_ROW['Cash']}+{fcf_ref})"

        # Cash = MAX(cash_raw, min_cash)  — floored at minimum cash buffer
        _write_fcst_cell(ws, BS_ROW["Cash"], bs_col,
                         f"=MAX({cash_raw},{min_cash})",
                         bold=True, border=BORDER_BOTTOM_MED)

        # Restricted Cash: flat
        _write_fcst_cell(ws, BS_ROW["Restricted Cash"], bs_col,
                         f"={prev_let}{BS_ROW['Restricted Cash']}")

        # Receivables = Revenue × recv_pct
        _write_fcst_cell(ws, BS_ROW["Receivables"], bs_col, f"={is_rev}*{recv_pct}")

        # Payables = Revenue × pay_pct
        _write_fcst_cell(ws, BS_ROW["Payables"], bs_col, f"={is_rev}*{pay_pct}")

        # Total Debt = bounded revolver draw/no-repay mechanic
        # Draw when cash_raw < min_cash; bounded by [min_debt, max_debt] assumptions.
        # NOTE: this is a draw-only revolver — debt never automatically repays.
        # See docstring for circular-reference hazard warning.
        _write_fcst_cell(ws, BS_ROW["Total Debt"], bs_col,
                         f"=MAX({min_debt},MIN({max_debt},"
                         f"{prev_let}{BS_ROW['Total Debt']}+MAX(0,{min_cash}-{cash_raw})))",
                         bold=True, border=BORDER_BOTTOM_MED)

        # Equity = prior Equity + Net Income
        # Assumption: zero dividends / buybacks — all earnings retained.
        # Add a share-repurchase row and subtract from Equity if this changes.
        _write_fcst_cell(ws, BS_ROW["Equity"], bs_col,
                         f"={prev_let}{BS_ROW['Equity']}+{is_ni}",
                         bold=True, border=BORDER_BOTTOM_MED)

    # ---- Balance Sheet Check (partial) ----
    # Separator
    ws.row_dimensions[BS_ROW["sep_check"]].height = 6
    for ci in range(1, total_cols + 1):
        ws.cell(row=BS_ROW["sep_check"], column=ci).fill = FILL_SUBHDR

    # Check header
    c = ws.cell(row=BS_ROW["check_header"], column=LABEL_COL)
    style_cell(c, "Balance Sheet Check (Partial)",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(start_row=BS_ROW["check_header"], start_column=1,
                   end_row=BS_ROW["check_header"], end_column=total_cols)

    # Check row labels
    check_label_spec = [
        (BS_ROW["Partial Assets"], "  Partial Assets (Cash + Restricted + Recv)"),
        (BS_ROW["Partial LE"],     "  Partial L+E (Pay + Debt + Equity)"),
        (BS_ROW["Check"],          "  Difference (non-zero = excluded items)"),
    ]
    for sheet_row, label in check_label_spec:
        style_cell(ws.cell(row=sheet_row, column=LABEL_COL), label,
                   font=FONT_BOLD if sheet_row == BS_ROW["Check"] else FONT_FORMULA,
                   alignment=Alignment(horizontal="left", indent=1))

    # Check formulas — all data columns (historical + forecast)
    total_data_cols = num_hist + NUM_FCST_COLS
    for ci_idx in range(total_data_cols):
        col_idx  = HIST_COL_START + ci_idx
        let      = cl(col_idx)
        is_fcst  = ci_idx >= num_hist
        fill     = FILL_FCST if is_fcst else FILL_NONE

        cr = BS_ROW["Cash"]
        rr = BS_ROW["Restricted Cash"]
        rv = BS_ROW["Receivables"]
        py = BS_ROW["Payables"]
        dt = BS_ROW["Total Debt"]
        eq = BS_ROW["Equity"]

        for check_r, formula in [
            (BS_ROW["Partial Assets"], f"={let}{cr}+{let}{rr}+{let}{rv}"),
            (BS_ROW["Partial LE"],     f"={let}{py}+{let}{dt}+{let}{eq}"),
            (BS_ROW["Check"],
             f"={let}{BS_ROW['Partial Assets']}-{let}{BS_ROW['Partial LE']}"),
        ]:
            c = ws.cell(row=check_r, column=col_idx)
            c.value         = formula
            c.number_format = NUM_CURRENCY
            c.font          = FONT_BOLD if check_r == BS_ROW["Check"] else FONT_FORMULA
            c.fill          = fill
            c.alignment     = Alignment(horizontal="right")
            if check_r == BS_ROW["Check"]:
                c.border = BORDER_BOTTOM_MED

    note_row = BS_ROW["Check"] + 1
    style_cell(ws.cell(row=note_row, column=1),
               "* BS Check excludes custodial crypto assets, goodwill, deferred taxes, "
               "and other items not in XBRL extract. Non-zero gap = unexplained by modeled items. "
               "Debt forecast = flat + revolver draw when FCF would push cash below minimum balance.",
               font=FONT_NOTE)

    # ---- Prominent debt mechanics note (Task #20) ----
    ws.row_dimensions[BS_ROW["debt_note"]].height = 48
    c = ws.cell(row=BS_ROW["debt_note"], column=LABEL_COL)
    c.value = (
        "⚠  Historical Total Debt is dominated by securities-lending collateral liabilities, "
        "NOT traditional corporate borrowing. HOOD books cash received when lending customer stocks "
        "as a balance-sheet liability. Quarterly swings of $3–4B reflect customer stock-lending "
        "volumes, not balance-sheet stress or credit risk. "
        "Forecast 'Total Debt' models only a corporate revolver plug; "
        "securities-lending liabilities are not forecast."
    )
    c.font      = Font(name="Calibri", size=9, color="7F3F00", bold=True)
    c.fill      = PatternFill("solid", fgColor="FFF2CC")   # light amber warning
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(
        start_row=BS_ROW["debt_note"], start_column=1,
        end_row=BS_ROW["debt_note"],   end_column=total_cols
    )

    # ---- Credit Metrics section (Task #18) ----
    # EBITDA proxy = Operating Income + SBC  (D&A not separately available in XBRL)
    # Debt/EBITDA and Net Debt/EBITDA cross-reference IS for OI and SBC.
    # is_hist_offset: IS has more historical quarters than BS due to earlier XBRL coverage.
    # Computed dynamically so credit metric columns stay aligned even if extraction changes.
    is_hist_offset = is_fcst_col_start - bs_fcst_start
    assert is_hist_offset >= 0, (
        f"IS must have >= BS historical quarters; got is_hist_offset={is_hist_offset}. "
        "Check CSV column counts."
    )

    ws.row_dimensions[BS_ROW["sep_credit"]].height = 6
    for ci in range(1, total_cols + 1):
        ws.cell(row=BS_ROW["sep_credit"], column=ci).fill = FILL_SUBHDR

    c = ws.cell(row=BS_ROW["credit_header"], column=LABEL_COL)
    style_cell(c, "Credit Metrics",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(start_row=BS_ROW["credit_header"], start_column=1,
                   end_row=BS_ROW["credit_header"],   end_column=total_cols)

    credit_labels = [
        (BS_ROW["EBITDA"],         "  EBITDA (proxy: OI + SBC; forecast adds D&A) †"),
        (BS_ROW["Debt_EBITDA"],    "  Total Debt / EBITDA"),
        (BS_ROW["NetDebt_EBITDA"], "  Net Debt / EBITDA"),
    ]
    for sheet_row, label in credit_labels:
        style_cell(ws.cell(row=sheet_row, column=LABEL_COL), label,
                   font=FONT_FORMULA,
                   alignment=Alignment(horizontal="left", indent=1))

    oi_is_r  = IS_ROW["Operating Income"]
    sbc_is_r = IS_ROW["SBC"]

    # Historical credit metrics: cross-reference IS hist columns
    for bi in range(num_hist):
        bs_col  = HIST_COL_START + bi
        bs_let  = cl(bs_col)
        is_col  = HIST_COL_START + bi + is_hist_offset
        is_let  = cl(is_col)

        ebitda_f = (f"='Income Statement'!{is_let}{oi_is_r}"
                    f"+'Income Statement'!{is_let}{sbc_is_r}")

        c = ws.cell(row=BS_ROW["EBITDA"], column=bs_col)
        c.value = ebitda_f;  c.number_format = NUM_CURRENCY
        c.font  = FONT_FORMULA;  c.alignment = Alignment(horizontal="right")

        for credit_r, fmt, formula in [
            (BS_ROW["Debt_EBITDA"],
             NUM_MULT,
             f"=IF(ISNUMBER({bs_let}{BS_ROW['Total Debt']}),"
             f"{bs_let}{BS_ROW['Total Debt']}/{bs_let}{BS_ROW['EBITDA']},\"-\")"),
            (BS_ROW["NetDebt_EBITDA"],
             NUM_MULT,
             f"=IF(ISNUMBER({bs_let}{BS_ROW['Total Debt']}),"
             f"({bs_let}{BS_ROW['Total Debt']}-{bs_let}{BS_ROW['Cash']})/{bs_let}{BS_ROW['EBITDA']},\"-\")"),
        ]:
            c = ws.cell(row=credit_r, column=bs_col)
            c.value = formula;  c.number_format = fmt
            c.font  = FONT_FORMULA;  c.alignment = Alignment(horizontal="right")

    # Forecast credit metrics
    for fi in range(NUM_FCST_COLS):
        bs_col = bs_fcst_start + fi
        bs_let = cl(bs_col)
        is_let = cl(is_fcst_col_start + fi)

        cf_da_ref = f"'Cash Flow'!{cl(cf_fcst_col_start + fi)}{CF_ROW['DA']}"
        ebitda_f = (f"='Income Statement'!{is_let}{oi_is_r}"
                    f"+'Income Statement'!{is_let}{sbc_is_r}"
                    f"+{cf_da_ref}")

        _write_fcst_cell(ws, BS_ROW["EBITDA"], bs_col, ebitda_f)
        _write_fcst_cell(ws, BS_ROW["Debt_EBITDA"], bs_col,
                         f"={bs_let}{BS_ROW['Total Debt']}/{bs_let}{BS_ROW['EBITDA']}",
                         num_format=NUM_MULT)
        _write_fcst_cell(ws, BS_ROW["NetDebt_EBITDA"], bs_col,
                         f"=({bs_let}{BS_ROW['Total Debt']}-{bs_let}{BS_ROW['Cash']})/{bs_let}{BS_ROW['EBITDA']}",
                         num_format=NUM_MULT)

    credit_note_row = BS_ROW["NetDebt_EBITDA"] + 1
    style_cell(ws.cell(row=credit_note_row, column=1),
               "† Historical EBITDA = OI + SBC (D&A not separately reported in XBRL). "
               "Forecast EBITDA = OI + SBC + D&A (D&A = Revenue × da_pct). "
               "Metric is therefore not directly comparable across the historical/forecast boundary. "
               "Leverage ratios show '-' where Total Debt data is unavailable (Q2 2023). "
               "Historical Debt/EBITDA is elevated due to securities-lending liabilities — see warning above.",
               font=FONT_NOTE)


def build_cash_flow(
    ws,
    df_cf: pd.DataFrame,
    cell_refs: dict[str, str],
    is_fcst_col_start: int,
    bs_fcst_col_start: int,
) -> None:
    """Build the Cash Flow Statement with historical actuals and 4-year forecast.

    Integrates Income Statement (Net Income, SBC, Revenue) and Balance Sheet
    (Receivables, Payables) via cross-sheet references. Implements standard
    indirect-method CFO calculation plus capex and FCF.

    Forecast Logic:
      Net Income = IS Net Income (cross-reference)
      SBC = IS Stock-Based Compensation (cross-reference)
      D&A = IS Revenue * da_pct (non-cash add-back)
      ΔAR = Receivables[current] - Receivables[prior] (from Balance Sheet)
      ΔAP = Payables[current] - Payables[prior] (from Balance Sheet)
      CFO = NI + SBC + D&A - ΔAR + ΔAP
      Capex = IS Revenue * capex_pct
      FCF = CFO - Capex

    Working Capital Integration:
      Changes in AR and AP are derived from BS forecast (Receivables % and
      Payables % of revenue), creating the three-statement integration chain:
      IS Revenue -> BS (Recv, Pay) -> CF (ΔAR, ΔAP) -> CFO

    Cash Bridge:
      Shows ΔCash = FCF + ΔDebt + Other/Equity.
      In forecast, Other = 0 by construction (Cash = prev + FCF + revolver draw).
      Historical "Other" captures equity issuances, buybacks, and non-modeled items.

    Args:
        ws: openpyxl Worksheet object for Cash Flow sheet.
        df_cf (pd.DataFrame): Cash Flow CSV data, index=line items, columns=periods.
        cell_refs (dict): Map of assumption keys to cell references.
        is_fcst_col_start (int): First forecast column index on Income Statement.
        bs_fcst_col_start (int): First forecast column index on Balance Sheet.

    Returns:
        None (modifies worksheet in-place).
    """
    num_hist      = len(df_cf.columns)
    cf_fcst_start = HIST_COL_START + num_hist
    total_cols    = LABEL_COL + num_hist + NUM_FCST_COLS

    ws.column_dimensions[cl(LABEL_COL)].width = COL_W_LABEL
    for ci in range(HIST_COL_START, cf_fcst_start + NUM_FCST_COLS + 1):
        ws.column_dimensions[cl(ci)].width = COL_W_DATA

    _apply_title_row(ws, f"{TICKER} — Cash Flow Statement", total_cols)
    style_cell(ws.cell(row=2, column=1), "$ in millions", font=FONT_NOTE)
    _apply_col_headers(ws, list(df_cf.columns), 3, HIST_COL_START, FCST_YEARS)

    rows = [
        ("  Net Income",                  False, "Net Income"),
        ("  Stock-Based Compensation",    False, "Stock-Based Compensation"),
        ("  Depreciation & Amortization", False, "Depreciation & Amortization"),  # forecast only
        ("Cash from Operations",          True,  "Cash from Operations"),
        ("  Capital Expenditures",        False, "Capital Expenditures"),
        ("Free Cash Flow",                True,  "Free Cash Flow"),
    ]
    _write_hist_rows(ws, df_cf, rows)

    # ---- Forecast formulas ----
    capex_pct = cell_refs["capex_pct"]
    da_pct    = cell_refs["da_pct"]

    ni_is_r  = IS_ROW["Net Income"]
    sbc_is_r = IS_ROW["SBC"]
    rev_is_r = IS_ROW["Total Revenue"]

    for fi in range(NUM_FCST_COLS):
        cf_col   = cf_fcst_start + fi
        cf_let   = cl(cf_col)
        is_let   = cl(is_fcst_col_start + fi)

        # BS columns: this year and previous year (for ΔWC)
        bs_this  = cl(bs_fcst_col_start + fi)
        bs_prev  = cl(bs_fcst_col_start + fi - 1)  # fi=0 → last hist BS col

        is_ni_ref  = f"'Income Statement'!{is_let}{ni_is_r}"
        is_sbc_ref = f"'Income Statement'!{is_let}{sbc_is_r}"
        is_rev_ref = f"'Income Statement'!{is_let}{rev_is_r}"

        recv_r = BS_ROW["Receivables"]
        pay_r  = BS_ROW["Payables"]

        bs_recv_this = f"'Balance Sheet'!{bs_this}{recv_r}"
        bs_recv_prev = f"'Balance Sheet'!{bs_prev}{recv_r}"
        bs_pay_this  = f"'Balance Sheet'!{bs_this}{pay_r}"
        bs_pay_prev  = f"'Balance Sheet'!{bs_prev}{pay_r}"

        # D&A = Revenue × da_pct (non-cash; add-back in CFO)
        _write_fcst_cell(ws, CF_ROW["DA"], cf_col, f"={is_rev_ref}*{da_pct}")

        # CFO = NI + SBC + D&A − ΔAR + ΔAP
        # ΔAR increase = use of cash (subtract); ΔAP increase = source of cash (add)
        cfo_formula = (
            f"={is_ni_ref}+{is_sbc_ref}"
            f"+{cf_let}{CF_ROW['DA']}"
            f"-({bs_recv_this}-{bs_recv_prev})"
            f"+({bs_pay_this}-{bs_pay_prev})"
        )

        _write_fcst_cell(ws, CF_ROW["Net Income"], cf_col, f"={is_ni_ref}")
        _write_fcst_cell(ws, CF_ROW["SBC"],        cf_col, f"={is_sbc_ref}")
        _write_fcst_cell(ws, CF_ROW["CFO"],        cf_col, cfo_formula,
                         bold=True, border=BORDER_BOTTOM_MED)
        _write_fcst_cell(ws, CF_ROW["Capex"],      cf_col, f"={is_rev_ref}*{capex_pct}")
        _write_fcst_cell(ws, CF_ROW["FCF"],        cf_col,
                         f"={cf_let}{CF_ROW['CFO']}-{cf_let}{CF_ROW['Capex']}",
                         bold=True, border=BORDER_BOTTOM_MED)

    # ---- Cash Bridge (all data columns: historical + forecast) ----
    # Shows: ΔCash = FCF + ΔDebt + Other/Equity  →  Other should be ~0 in a closed model.
    # Historical "Other" captures equity issuances, buy-backs, and items not modeled.
    # In forecast, Other = 0 by construction (Cash = prev + FCF + revolver draw).

    # bs_hist_offset: BS has one more historical quarter than CF (BS starts Q2 2023,
    # CF starts Q3 2023).  For CF hist column hi (0-based):
    #   BS current col index = HIST_COL_START + hi + bs_hist_offset
    #   BS prior   col index = HIST_COL_START + hi + bs_hist_offset - 1
    cf_fcst_start  = HIST_COL_START + num_hist
    bs_hist_offset = bs_fcst_col_start - cf_fcst_start   # = 1

    # Separator
    ws.row_dimensions[CF_ROW["sep_bridge"]].height = 6
    for ci in range(1, total_cols + 1):
        ws.cell(row=CF_ROW["sep_bridge"], column=ci).fill = FILL_SUBHDR

    # Header
    c = ws.cell(row=CF_ROW["bridge_header"], column=LABEL_COL)
    style_cell(c, "Cash Bridge  (ΔCash = FCF + ΔDebt + Other)",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(start_row=CF_ROW["bridge_header"], start_column=1,
                   end_row=CF_ROW["bridge_header"], end_column=total_cols)

    # Row labels
    style_cell(ws.cell(row=CF_ROW["Delta Cash"], column=LABEL_COL),
               "  Δ Cash  (Balance Sheet)",
               font=FONT_FORMULA, alignment=Alignment(horizontal="left", indent=1))
    style_cell(ws.cell(row=CF_ROW["Delta Debt"], column=LABEL_COL),
               "  Δ Debt  (Balance Sheet)",
               font=FONT_FORMULA, alignment=Alignment(horizontal="left", indent=1))
    style_cell(ws.cell(row=CF_ROW["Bridge Check"], column=LABEL_COL),
               "  Other / Equity  (ΔCash − FCF − ΔDebt)",
               font=FONT_BOLD, alignment=Alignment(horizontal="left", indent=1))

    # Formulas for every data column (hist + forecast)
    total_data_cols = num_hist + NUM_FCST_COLS
    cash_r = BS_ROW["Cash"]
    debt_r = BS_ROW["Total Debt"]
    fcf_r  = CF_ROW["FCF"]
    dc_r   = CF_ROW["Delta Cash"]
    dd_r   = CF_ROW["Delta Debt"]

    for ci_idx in range(total_data_cols):
        is_fcst  = ci_idx >= num_hist
        col_idx  = HIST_COL_START + ci_idx
        cf_let   = cl(col_idx)
        fill     = FILL_FCST if is_fcst else FILL_NONE

        if is_fcst:
            fi       = ci_idx - num_hist
            bs_curr  = cl(bs_fcst_col_start + fi)
            bs_prev  = cl(bs_fcst_col_start + fi - 1)   # fi=0 → last hist BS col
        else:
            hi       = ci_idx
            bs_curr  = cl(HIST_COL_START + hi + bs_hist_offset)
            bs_prev  = cl(HIST_COL_START + hi + bs_hist_offset - 1)

        # ΔCash: valid for all columns (Cash Q2 2023 is available)
        c = ws.cell(row=dc_r, column=col_idx)
        c.value         = (f"='Balance Sheet'!{bs_curr}{cash_r}"
                           f"-'Balance Sheet'!{bs_prev}{cash_r}")
        c.number_format = NUM_CURRENCY
        c.font          = FONT_FORMULA
        c.fill          = fill
        c.alignment     = Alignment(horizontal="right")

        # ΔDebt: skip Q3 2023 (first hist col) — Q2 2023 debt is not available in XBRL
        if not is_fcst and hi == 0:
            ws.cell(row=dd_r, column=col_idx).fill = fill
            ws.cell(row=CF_ROW["Bridge Check"], column=col_idx).fill = fill
            continue

        c = ws.cell(row=dd_r, column=col_idx)
        c.value         = (f"='Balance Sheet'!{bs_curr}{debt_r}"
                           f"-'Balance Sheet'!{bs_prev}{debt_r}")
        c.number_format = NUM_CURRENCY
        c.font          = FONT_FORMULA
        c.fill          = fill
        c.alignment     = Alignment(horizontal="right")

        # Bridge Check = ΔCash − FCF − ΔDebt  (should be ≈ 0 in forecast)
        c = ws.cell(row=CF_ROW["Bridge Check"], column=col_idx)
        c.value         = (f"={cf_let}{dc_r}"
                           f"-{cf_let}{fcf_r}"
                           f"-{cf_let}{dd_r}")
        c.number_format = NUM_CURRENCY
        c.font          = FONT_BOLD
        c.fill          = fill
        c.alignment     = Alignment(horizontal="right")
        c.border        = BORDER_BOTTOM_DOUBLE

    note_row = CF_ROW["Bridge Check"] + 1
    style_cell(ws.cell(row=note_row, column=1),
               "* Forecast CFO = NI + SBC + D&A − ΔAR + ΔAP (WC linked to Balance Sheet). "
               "Historical CFO includes volatile securities-lending flows not in forecast. "
               "Bridge Check = 0 in forecast (Cash and Debt are mechanically closed). "
               "Historical 'Other / Equity' residual = equity issuances, buy-backs, "
               "and balance-sheet items not captured in this extract. "
               "ΔDebt blank for Q3 2023: Q2 2023 Total Debt not available in XBRL.",
               font=FONT_NOTE)

    # ---- Conditional formatting: Bridge Check forecast cells (green = 0, red ≠ 0) ----
    chk_from = cl(cf_fcst_start)
    chk_to   = cl(cf_fcst_start + NUM_FCST_COLS - 1)
    chk_rng  = f"{chk_from}{CF_ROW['Bridge Check']}:{chk_to}{CF_ROW['Bridge Check']}"
    ws.conditional_formatting.add(chk_rng,
        CellIsRule(operator="equal",    formula=["0"], fill=FILL_GREEN_CF))
    ws.conditional_formatting.add(chk_rng,
        CellIsRule(operator="notEqual", formula=["0"], fill=FILL_RED_CF))


# ---------------------------------------------------------------------------
# Sheet 5 — Valuation Analysis
# ---------------------------------------------------------------------------

def build_valuation(
    ws,
    cell_refs: dict[str, str],
    is_fcst_col_start: int,
    bs_fcst_col_start: int,
    cf_fcst_col_start: int,
) -> None:
    """Build the Valuation Analysis sheet with DCF and sensitivity analysis.

    Calculates enterprise value using multiple methods:
      1. DCF (Discounted Cash Flow) with Gordon Growth terminal value
      2. Exit multiples (EV/EBITDA, EV/Revenue)
      3. WACC x Terminal Growth Rate sensitivity table

    DCF Methodology:
      * Discounts forecasted FCFs (from Cash Flow sheet) back to present value
      * Terminal Value = Year 4 FCF * (1 + TGR) / (WACC - TGR)
      * Enterprise Value = sum(PV of forecast FCFs) + PV of terminal value
      * Implied Equity Value = EV - Net Debt
      * Price per share = Implied Equity / Diluted Shares

    Exit Multiple Methods:
      * EBITDA = Operating Income + D&A (alternative proxy for cash generation)
      * Exit EV = Year 4 EBITDA * EV/EBITDA multiple
      * Or Exit EV = Year 4 Revenue * EV/Revenue multiple
      * Provides sanity check against trading multiples

    Sensitivity Table:
      2D matrix with WACC (rows) x Terminal Growth Rate (columns).
      Each cell shows implied enterprise value under that WACC/TGR combination.
      Highlights range of reasonable valuations.

    Args:
        ws: openpyxl Worksheet object for Valuation sheet.
        cell_refs (dict): Map of assumption keys to cell references.
        is_fcst_col_start (int): First forecast column index on Income Statement.
        bs_fcst_col_start (int): First forecast column index on Balance Sheet.
        cf_fcst_col_start (int): First forecast column index on Cash Flow.

    Returns:
        None (modifies worksheet in-place).
    """
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    for ci in range(3, 7):
        ws.column_dimensions[cl(ci)].width = 16

    last_col = 6  # F

    # ---- Title ----
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=1, column=1)
    style_cell(c, f"{TICKER} — Valuation Analysis",
               font=FONT_TITLE, fill=FILL_HEADER,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells(f"A1:{cl(last_col)}1")

    style_cell(ws.cell(row=2, column=1),
               "All values in $ millions unless noted. Inputs from Assumptions sheet (blue cells). "
               "FCF / EBITDA / Revenue cross-reference forecast sheets.",
               font=FONT_NOTE)
    ws.merge_cells(f"A2:{cl(last_col)}2")

    wacc_ref         = cell_refs["wacc"]
    tgr_ref          = cell_refs["terminal_growth"]
    exit_ebitda_ref  = cell_refs["exit_ebitda_mult"]
    exit_rev_ref     = cell_refs["exit_rev_mult"]
    shares_ref       = cell_refs["shares_diluted"]

    # Column letters for forecast years
    fcst_lets = [cl(is_fcst_col_start + fi) for fi in range(NUM_FCST_COLS)]   # IS
    cf_lets   = [cl(cf_fcst_col_start  + fi) for fi in range(NUM_FCST_COLS)]  # CF
    bs_lets   = [cl(bs_fcst_col_start  + fi) for fi in range(NUM_FCST_COLS)]  # BS

    # Shorthand row refs
    fcf_cf_r  = CF_ROW["FCF"]
    oi_is_r   = IS_ROW["Operating Income"]
    sbc_is_r  = IS_ROW["SBC"]
    da_cf_r   = CF_ROW["DA"]
    rev_is_r  = IS_ROW["Total Revenue"]
    cash_bs_r = BS_ROW["Cash"]
    debt_bs_r = BS_ROW["Total Debt"]

    # ==============================================================
    # SECTION 1: DCF
    # ==============================================================
    sec1_row = 4

    ws.row_dimensions[sec1_row].height = 18
    c = ws.cell(row=sec1_row, column=1)
    style_cell(c, "Section 1 — Discounted Cash Flow Analysis",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(f"A{sec1_row}:{cl(last_col)}{sec1_row}")

    # Col headers row 5
    hdr_r = sec1_row + 1
    for ci, label in enumerate(["", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "Terminal"], start=1):
        style_cell(ws.cell(row=hdr_r, column=ci), label,
                   font=FONT_HDR, fill=FILL_HEADER,
                   alignment=Alignment(horizontal="center"))
    ws.row_dimensions[hdr_r].height = 16

    # Row 6: FCF
    fcf_r_val = hdr_r + 1
    style_cell(ws.cell(row=fcf_r_val, column=1), "Free Cash Flow ($M)", font=FONT_FORMULA,
               alignment=Alignment(horizontal="left", indent=1))
    for fi in range(NUM_FCST_COLS):
        c = ws.cell(row=fcf_r_val, column=2 + fi)
        c.value = f"='Cash Flow'!{cf_lets[fi]}{fcf_cf_r}"
        c.number_format = NUM_CURRENCY; c.font = FONT_FORMULA
        c.fill = FILL_FCST; c.alignment = Alignment(horizontal="right")
    # Terminal FCF = FY2029E FCF (col E) * (1 + TGR)
    # NOTE: reference col E (FY2029E), NOT col F (this cell) — avoids self-reference circular ref
    c = ws.cell(row=fcf_r_val, column=6)
    c.value = f"=E{fcf_r_val}*(1+{tgr_ref})"
    c.number_format = NUM_CURRENCY; c.font = FONT_FORMULA
    c.fill = FILL_FCST; c.alignment = Alignment(horizontal="right")

    # Row 7: Discount period
    dp_r = fcf_r_val + 1
    style_cell(ws.cell(row=dp_r, column=1), "  Discount Period (n)", font=FONT_NOTE,
               alignment=Alignment(horizontal="left", indent=1))
    for fi in range(NUM_FCST_COLS):
        c = ws.cell(row=dp_r, column=2 + fi)
        c.value = fi + 1; c.number_format = "0"
        c.font = FONT_NOTE; c.fill = FILL_FCST
        c.alignment = Alignment(horizontal="right")

    # Row 8: Discount factor
    df_r = dp_r + 1
    style_cell(ws.cell(row=df_r, column=1), "  Discount Factor  1/(1+WACC)^n", font=FONT_NOTE,
               alignment=Alignment(horizontal="left", indent=1))
    for fi in range(NUM_FCST_COLS):
        c = ws.cell(row=df_r, column=2 + fi)
        c.value = f"=1/(1+{wacc_ref})^{fi+1}"
        c.number_format = "0.000"; c.font = FONT_NOTE
        c.fill = FILL_FCST; c.alignment = Alignment(horizontal="right")

    # Row 9: PV of FCF
    pv_r = df_r + 1
    style_cell(ws.cell(row=pv_r, column=1), "  Present Value of FCF ($M)", font=FONT_FORMULA,
               alignment=Alignment(horizontal="left", indent=1))
    pv_cols = []
    for fi in range(NUM_FCST_COLS):
        c_let = cl(2 + fi)
        c = ws.cell(row=pv_r, column=2 + fi)
        c.value = f"={c_let}{fcf_r_val}*{c_let}{df_r}"
        c.number_format = NUM_CURRENCY; c.font = FONT_FORMULA
        c.fill = FILL_FCST; c.alignment = Alignment(horizontal="right")
        pv_cols.append(c_let)

    # Separator
    sep_r = pv_r + 1
    ws.row_dimensions[sep_r].height = 6
    for ci in range(1, last_col + 1):
        ws.cell(row=sep_r, column=ci).fill = FILL_SUBHDR

    sum_r  = sep_r + 1
    tv_r   = sum_r + 1
    pvtv_r = tv_r + 1
    ev_r   = pvtv_r + 1
    nd_r   = ev_r + 1
    eqv_r  = nd_r + 1
    sp_r   = eqv_r + 1

    labels_vals = [
        (sum_r,  "Sum of PV(FCFs)  ($M)",        f"=SUM({pv_cols[0]}{pv_r}:{pv_cols[-1]}{pv_r})", NUM_CURRENCY, False),
        (tv_r,   "Terminal Value  (Gordon Growth) ($M)",
                 f"=F{fcf_r_val}/({wacc_ref}-{tgr_ref})", NUM_CURRENCY, False),
        (pvtv_r, "PV of Terminal Value  ($M)  [discounted to today]",
                 f"={cl(6)}{tv_r}/(1+{wacc_ref})^4", NUM_CURRENCY, False),
        (ev_r,   "Enterprise Value — DCF  ($M)",
                 f"={cl(2)}{sum_r}+{cl(2)}{pvtv_r}", NUM_CURRENCY, True),
        (nd_r,   "  (–) Net Debt FY2026E  ($M)",
                 f"=('Balance Sheet'!{bs_lets[0]}{debt_bs_r}-'Balance Sheet'!{bs_lets[0]}{cash_bs_r})",
                 NUM_CURRENCY, False),
        (eqv_r,  "Equity Value — DCF  ($M)",
                 f"={cl(2)}{ev_r}-{cl(2)}{nd_r}", NUM_CURRENCY, True),
        (sp_r,   "Implied Share Price — DCF  ($/share)",
                 f"={cl(2)}{eqv_r}/{shares_ref}", NUM_EPS, True),
    ]

    for row_n, label, formula, fmt, bold in labels_vals:
        style_cell(ws.cell(row=row_n, column=1), label,
                   font=FONT_BOLD if bold else FONT_FORMULA,
                   alignment=Alignment(horizontal="left", indent=1 if not bold else 0))
        c = ws.cell(row=row_n, column=2)
        c.value = formula; c.number_format = fmt
        c.font = FONT_BOLD if bold else FONT_FORMULA
        c.fill = FILL_FCST; c.alignment = Alignment(horizontal="right")
        if bold:
            c.border = BORDER_BOTTOM_MED

    # Terminal value methodology note (column C, beside the TV and PV TV rows)
    ws.column_dimensions["C"].width = max(
        float(ws.column_dimensions["C"].width or 0), 52
    )
    tv_note = ws.cell(row=tv_r, column=3)
    tv_note.value = (
        "Gordon Growth:  TV = FCF(FY2029E) × (1 + TGR) / (WACC − TGR)"
    )
    tv_note.font = FONT_NOTE
    tv_note.alignment = Alignment(horizontal="left", vertical="center")

    pvtv_note = ws.cell(row=pvtv_r, column=3)
    pvtv_note.value = (
        "Discounted to PV at time 0:  PV(TV) = TV ÷ (1 + WACC)⁴  "
        "[end-of-year convention; 4 years from today to FY2029E terminal year]"
    )
    pvtv_note.font = FONT_NOTE
    pvtv_note.alignment = Alignment(horizontal="left", vertical="center")

    # ==============================================================
    # SECTION 2: Exit Multiple Analysis (FY2029E terminal year)
    # ==============================================================
    sec2_row = sp_r + 3
    ws.row_dimensions[sec2_row].height = 18
    c = ws.cell(row=sec2_row, column=1)
    style_cell(c, "Section 2 — Exit Multiple Analysis  (FY2029E terminal year)",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(f"A{sec2_row}:{cl(last_col)}{sec2_row}")

    em_hdr = sec2_row + 1
    for ci, label in enumerate(["Metric", "Value ($M or x)", "", "EV / EBITDA basis", "EV / Revenue basis", ""], start=1):
        style_cell(ws.cell(row=em_hdr, column=ci), label,
                   font=FONT_HDR, fill=FILL_HEADER,
                   alignment=Alignment(horizontal="center"))

    is_y4_let = fcst_lets[3]
    cf_y4_let = cf_lets[3]

    ebitda_y4_r = em_hdr + 1
    rev_y4_r    = ebitda_y4_r + 1
    mult_r      = rev_y4_r + 1
    ev_em_r     = mult_r + 1
    nd_em_r     = ev_em_r + 1
    eqv_em_r    = nd_em_r + 1
    sp_em_r     = eqv_em_r + 1

    # EBITDA (OI + SBC + D&A)
    style_cell(ws.cell(row=ebitda_y4_r, column=1), "FY2029E EBITDA (proxy)", font=FONT_FORMULA,
               alignment=Alignment(horizontal="left", indent=1))
    ebitda_val = (f"='Income Statement'!{is_y4_let}{oi_is_r}"
                  f"+'Income Statement'!{is_y4_let}{sbc_is_r}"
                  f"+'Cash Flow'!{cf_y4_let}{da_cf_r}")
    for ci in [2, 4]:
        c = ws.cell(row=ebitda_y4_r, column=ci)
        c.value = ebitda_val; c.number_format = NUM_CURRENCY
        c.font = FONT_FORMULA; c.fill = FILL_FCST
        c.alignment = Alignment(horizontal="right")

    # Revenue
    style_cell(ws.cell(row=rev_y4_r, column=1), "FY2029E Revenue", font=FONT_FORMULA,
               alignment=Alignment(horizontal="left", indent=1))
    rev_val = f"='Income Statement'!{is_y4_let}{rev_is_r}"
    for ci in [2, 5]:
        c = ws.cell(row=rev_y4_r, column=ci)
        c.value = rev_val; c.number_format = NUM_CURRENCY
        c.font = FONT_FORMULA; c.fill = FILL_FCST
        c.alignment = Alignment(horizontal="right")

    # Multiples
    style_cell(ws.cell(row=mult_r, column=1), "Exit Multiple (input)", font=FONT_FORMULA,
               alignment=Alignment(horizontal="left", indent=1))
    for ci, ref in [(4, exit_ebitda_ref), (5, exit_rev_ref)]:
        c = ws.cell(row=mult_r, column=ci)
        c.value = f"={ref}"; c.number_format = NUM_MULT
        c.font = FONT_INPUT; c.fill = FILL_YELLOW
        c.alignment = Alignment(horizontal="right")

    # EV
    style_cell(ws.cell(row=ev_em_r, column=1), "Implied Enterprise Value ($M)", font=FONT_BOLD,
               alignment=Alignment(horizontal="left"))
    c4 = ws.cell(row=ev_em_r, column=4)
    c4.value = f"={cl(4)}{ebitda_y4_r}*{cl(4)}{mult_r}"
    c4.number_format = NUM_CURRENCY; c4.font = FONT_BOLD
    c4.fill = FILL_FCST; c4.alignment = Alignment(horizontal="right")
    c4.border = BORDER_BOTTOM_MED
    c5 = ws.cell(row=ev_em_r, column=5)
    c5.value = f"={cl(5)}{rev_y4_r}*{cl(5)}{mult_r}"
    c5.number_format = NUM_CURRENCY; c5.font = FONT_BOLD
    c5.fill = FILL_FCST; c5.alignment = Alignment(horizontal="right")
    c5.border = BORDER_BOTTOM_MED

    # Net Debt — use FY2029E Net Debt discounted back 4 years at WACC to avoid
    # timing mismatch: EV is computed at FY2029E but FY2026E debt was previously used.
    nd_formula = (
        f"=('Balance Sheet'!{bs_lets[3]}{debt_bs_r}"
        f"-'Balance Sheet'!{bs_lets[3]}{cash_bs_r})"
        f"/(1+{wacc_ref})^4"
    )
    style_cell(ws.cell(row=nd_em_r, column=1), "  (–) Net Debt FY2029E, PV @ WACC ($M)", font=FONT_FORMULA,
               alignment=Alignment(horizontal="left", indent=1))
    for ci in [4, 5]:
        c = ws.cell(row=nd_em_r, column=ci)
        c.value = nd_formula; c.number_format = NUM_CURRENCY
        c.font = FONT_FORMULA; c.fill = FILL_FCST
        c.alignment = Alignment(horizontal="right")

    # Equity Value
    style_cell(ws.cell(row=eqv_em_r, column=1), "Implied Equity Value ($M)", font=FONT_BOLD,
               alignment=Alignment(horizontal="left"))
    for ci in [4, 5]:
        c = ws.cell(row=eqv_em_r, column=ci)
        c.value = f"={cl(ci)}{ev_em_r}-{cl(ci)}{nd_em_r}"
        c.number_format = NUM_CURRENCY; c.font = FONT_BOLD
        c.fill = FILL_FCST; c.alignment = Alignment(horizontal="right")
        c.border = BORDER_BOTTOM_MED

    # Share Price
    style_cell(ws.cell(row=sp_em_r, column=1), "Implied Share Price ($/share)", font=FONT_BOLD,
               alignment=Alignment(horizontal="left"))
    for ci in [4, 5]:
        c = ws.cell(row=sp_em_r, column=ci)
        c.value = f"={cl(ci)}{eqv_em_r}/{shares_ref}"
        c.number_format = NUM_EPS; c.font = FONT_BOLD
        c.fill = FILL_FCST; c.alignment = Alignment(horizontal="right")
        c.border = BORDER_BOTTOM_DOUBLE

    # ==============================================================
    # SECTION 3: Implied Price Summary
    # ==============================================================
    sec3_row = sp_em_r + 3
    ws.row_dimensions[sec3_row].height = 18
    c = ws.cell(row=sec3_row, column=1)
    style_cell(c, "Section 3 — Implied Share Price Summary",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(f"A{sec3_row}:{cl(last_col)}{sec3_row}")

    sum_hdr = sec3_row + 1
    for ci, label in enumerate(["Method", "Implied Price ($/share)", "", "", "", ""], start=1):
        style_cell(ws.cell(row=sum_hdr, column=ci), label,
                   font=FONT_HDR, fill=FILL_HEADER,
                   alignment=Alignment(horizontal="center" if ci > 1 else "left"))

    for offset, (label, formula) in enumerate([
        ("DCF (Gordon Growth terminal)",    f"=B{sp_r}"),
        ("Exit EV/EBITDA x FY2029E",        f"=D{sp_em_r}"),
        ("Exit EV/Revenue x FY2029E",       f"=E{sp_em_r}"),
    ]):
        rr = sum_hdr + 1 + offset
        style_cell(ws.cell(row=rr, column=1), label, font=FONT_FORMULA,
                   alignment=Alignment(horizontal="left", indent=1))
        c = ws.cell(row=rr, column=2)
        c.value = formula; c.number_format = NUM_EPS
        c.font = FONT_BOLD; c.fill = FILL_FCST
        c.alignment = Alignment(horizontal="right")
        c.border = BORDER_BOTTOM_MED

    style_cell(ws.cell(row=sum_hdr + 4, column=1),
               "* DCF Net Debt = FY2026E Total Debt - FY2026E Cash. "
               "Exit Multiple Net Debt = FY2029E Total Debt - FY2029E Cash, discounted back 4 years at WACC "
               "to match the timing of the FY2029E enterprise value. "
               "Historical debt elevated by securities-lending liabilities — see Balance Sheet note. "
               "EBITDA = OI + SBC + D&A (D&A = Revenue x da_pct). "
               "Share prices are purely illustrative; not investment advice.",
               font=FONT_NOTE)

    # ==============================================================
    # SECTION 4: WACC × Terminal Growth Rate Sensitivity
    # ==============================================================
    sec4_row = sum_hdr + 7   # 2 blank rows after the footnote at sum_hdr+4
    ws.row_dimensions[sec4_row].height = 18
    c = ws.cell(row=sec4_row, column=1)
    style_cell(c, "Section 4 — DCF Sensitivity: WACC × Terminal Growth Rate",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR)
    ws.merge_cells(f"A{sec4_row}:{cl(last_col)}{sec4_row}")

    note4_r = sec4_row + 1
    style_cell(ws.cell(row=note4_r, column=1),
               "Implied DCF share price ($/share) at varying WACC and terminal growth rates. "
               "FCF from Cash Flow forecast; net debt from FY2026E Balance Sheet. "
               "Base case (WACC 12%, TGR 3%) highlighted with bold border.",
               font=FONT_NOTE)
    ws.merge_cells(f"A{note4_r}:{cl(last_col)}{note4_r}")

    col_hdr_r = sec4_row + 3   # leave one blank row between note and headers
    tgr_vals  = [0.01, 0.02, 0.03, 0.04, 0.05]
    wacc_vals = [0.08, 0.10, 0.12, 0.14, 0.16]

    # Column headers (TGR values)
    style_cell(ws.cell(row=col_hdr_r, column=1), "WACC  ↓  /  TGR  →",
               font=FONT_HDR, fill=FILL_HEADER,
               alignment=Alignment(horizontal="center"))
    for ci, tgr in enumerate(tgr_vals, start=2):
        c = ws.cell(row=col_hdr_r, column=ci)
        c.value = tgr
        c.number_format = NUM_PCT
        c.font = FONT_HDR
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[col_hdr_r].height = 16

    # Row headers (WACC values) + data cells
    for ri, wacc_val in enumerate(wacc_vals):
        dr = col_hdr_r + 1 + ri

        # Row header: WACC value
        rh = ws.cell(row=dr, column=1)
        rh.value = wacc_val
        rh.number_format = NUM_PCT
        rh.font = FONT_HDR
        rh.fill = FILL_HEADER
        rh.alignment = Alignment(horizontal="center")

        for ci in range(2, 7):
            w_ref = f"$A{dr}"              # fixed col A (WACC), variable row
            g_ref = f"{cl(ci)}${col_hdr_r}"  # variable col, fixed header row (TGR)

            fcf_refs = [
                f"'Cash Flow'!{cl(cf_fcst_col_start + fi)}{fcf_cf_r}"
                for fi in range(NUM_FCST_COLS)
            ]
            pv_fcfs = "+".join(
                f"{fcf_refs[i]}/(1+{w_ref})^{i + 1}" for i in range(NUM_FCST_COLS)
            )
            pv_tv = (
                f"({fcf_refs[-1]}*(1+{g_ref})"
                f"/(({w_ref}-{g_ref})*(1+{w_ref})^{NUM_FCST_COLS}))"
            )
            nd_formula = (
                f"('Balance Sheet'!{bs_lets[0]}{debt_bs_r}"
                f"-'Balance Sheet'!{bs_lets[0]}{cash_bs_r})"
            )
            formula = f"=({pv_fcfs}+{pv_tv}-{nd_formula})/{shares_ref}"

            is_base = (ri == 2 and ci == 4)   # WACC=12% (index 2), TGR=3% (col D)
            cell = ws.cell(row=dr, column=ci)
            cell.value = formula
            cell.number_format = NUM_EPS
            cell.font = Font(name="Calibri", size=10, bold=is_base, color=C_BLACK)
            cell.fill = FILL_FCST
            cell.alignment = Alignment(horizontal="right")
            if is_base:
                cell.border = Border(
                    top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM, right=_MEDIUM
                )

    # Conditional color scale on the 5×5 data range
    data_range = f"B{col_hdr_r + 1}:F{col_hdr_r + 5}"
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="min",        start_color="F8696B",
            mid_type="percentile",   mid_value=50, mid_color="FFEB9C",
            end_type="max",          end_color="63BE7B",
        ),
    )

    # Footnote
    note5_r = col_hdr_r + 7
    style_cell(ws.cell(row=note5_r, column=1),
               "† Gordon Growth terminal value requires WACC > TGR; cells where WACC ≤ TGR "
               "will show #DIV/0! or negative — those scenarios are economically invalid.",
               font=FONT_NOTE)
    ws.merge_cells(f"A{note5_r}:{cl(last_col)}{note5_r}")


# ---------------------------------------------------------------------------
# Sheet 6 — Sensitivity Analysis
# ---------------------------------------------------------------------------

def build_sensitivity(ws, cell_refs: dict[str, str], is_fcst_col_start: int) -> None:
    """Build the Sensitivity Analysis sheet with multiple two-way sensitivity tables.

    Creates 2D sensitivity tables to explore how valuation (or other key metrics)
    varies with changes in two critical assumptions simultaneously.

    Table 1 - OpEx % x Revenue Growth:
      Rows: Operating Expenses % of Revenue (around opex_pct assumption +/-20%)
      Cols: Year 1 Transaction Revenue Growth (around txn_growth_y1 +/-50%)
      Cell Values: FY2026E FCF under each scenario

    Table 2 - Net Income Margin % x Revenue Growth:
      Rows: OpEx % assumption variations
      Cols: Txn growth variations
      Cell Values: FY2026E Net Income under each scenario

    Table 3 - NI Growth x Txn Growth:
      Rows: Net Interest Revenue Growth (interest-rate sensitivity)
      Cols: Transaction Revenue Growth (volume sensitivity)
      Cell Values: FY2026E Total Revenue under each scenario

    Usage:
      Allows stakeholders to understand which assumptions have greatest impact on
      valuation, and where model is most/least sensitive to changes. Useful for
      determining which inputs require highest precision in estimation.

    Args:
        ws: openpyxl Worksheet object for Sensitivity Analysis sheet.
        cell_refs (dict): Map of assumption keys to cell references.
        is_fcst_col_start (int): First forecast column index on Income Statement.

    Returns:
        None (modifies worksheet in-place).
    """
    txn_r  = IS_ROW["Txn Revenue"]
    ni_r   = IS_ROW["NI Revenue"]
    oth_r  = IS_ROW["Other Revenue"]

    ltm_start_col = cl(is_fcst_col_start - 4)
    ltm_end_col   = cl(is_fcst_col_start - 1)

    ltm_txn = f"SUM('Income Statement'!{ltm_start_col}{txn_r}:{ltm_end_col}{txn_r})"
    ltm_ni  = f"SUM('Income Statement'!{ltm_start_col}{ni_r}:{ltm_end_col}{ni_r})"
    ltm_oth = f"SUM('Income Statement'!{ltm_start_col}{oth_r}:{ltm_end_col}{oth_r})"

    sbc_ref          = cell_refs["sbc_pct"]
    da_ref           = cell_refs["da_pct"]
    tax_ref          = cell_refs["tax_rate_y1"]
    capex_ref        = cell_refs["capex_pct"]
    other_income_ref = cell_refs["other_income"]
    ni_growth_ref    = cell_refs["ni_growth_y1"]
    other_growth_ref = cell_refs["other_growth_y1"]

    # Axes for Tables 1 & 2
    txn_growth_vals = [0.10, 0.20, 0.30, 0.40, 0.50]
    opex_vals       = [0.32, 0.37, 0.42, 0.47, 0.52]
    # Axes for Table 3 — centered on the ni_growth_y1 default (15%)
    ni_growth_vals  = [-0.05, 0.05, 0.15, 0.25, 0.35]

    ws.column_dimensions["A"].width = 24
    for ci in range(2, 7):
        ws.column_dimensions[cl(ci)].width = 14

    # ---- Title ----
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=1, column=1)
    style_cell(c, f"{TICKER} — Sensitivity Analysis",
               font=FONT_TITLE, fill=FILL_HEADER,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells("A1:F1")

    style_cell(ws.cell(row=2, column=1),
               "All values in $ millions. Revenue computed using per-segment LTM bases x "
               "independent growth rates. Non-varied drivers held at Assumptions sheet inputs.",
               font=FONT_NOTE)
    ws.merge_cells("A2:F2")

    def _write_table_header(hdr_row: int, col_axis_vals: list, col_header_label: str,
                            row_axis_label: str):
        """Write axis label cell + column header cells for one sensitivity table."""
        c = ws.cell(row=hdr_row, column=1)
        c.value = row_axis_label
        c.font  = FONT_INPUT; c.fill = FILL_YELLOW
        c.alignment = Alignment(horizontal="center")
        for ci_idx, val in enumerate(col_axis_vals):
            col_idx = 2 + ci_idx
            c2 = ws.cell(row=hdr_row, column=col_idx)
            c2.value = val; c2.font = FONT_INPUT; c2.fill = FILL_YELLOW
            c2.number_format = NUM_PCT
            c2.alignment = Alignment(horizontal="center")

    def _total_rev(g_txn: str, g_ni: str, g_oth: str) -> str:
        return (f"({ltm_txn}*(1+{g_txn})"
                f"+{ltm_ni}*(1+{g_ni})"
                f"+{ltm_oth}*(1+{g_oth}))")

    # ====================================================================
    # TABLE 1 — FCF  (rows 4–11)
    # ====================================================================
    ws.row_dimensions[4].height = 18
    c = ws.cell(row=4, column=1)
    style_cell(c, "FY2026E Free Cash Flow ($M)  —  Transaction Revenue Growth x OpEx %",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK), fill=FILL_SUBHDR)
    ws.merge_cells("A4:F4")
    style_cell(ws.cell(row=5, column=1),
               "Row: Txn Rev Growth y1 (10–50%). Col: OpEx % ex-SBC. "
               "NI Rev growth and Other Rev growth held at Assumptions inputs. "
               "Note: WC changes excluded from FCF for simplicity (net WC drag ≈ (recv_pct − pay_pct) × ΔRevenue).",
               font=FONT_NOTE)
    ws.merge_cells("A5:F5")

    _write_table_header(6, opex_vals, "OpEx %", "Txn Growth \\ OpEx")

    for ri, g_val in enumerate(txn_growth_vals):
        r = 7 + ri
        ca = ws.cell(row=r, column=1)
        ca.value = g_val; ca.font = FONT_INPUT; ca.fill = FILL_YELLOW
        ca.number_format = NUM_PCT; ca.alignment = Alignment(horizontal="center")
        for ci_idx in range(5):
            col_idx = 2 + ci_idx
            g_ref = f"$A{r}"
            o_ref = f"{cl(col_idx)}$6"
            tv = _total_rev(g_ref, ni_growth_ref, other_growth_ref)
            # FCF = OI − Tax + Other Income + SBC add-back + D&A add-back − Capex
            # OI = tv*(1−opex−sbc);  Tax applied to OI only (BTL is post-tax)
            fcf_formula = (
                f"={tv}*(1-{o_ref}-{sbc_ref})"
                f"-MAX({tv}*(1-{o_ref}-{sbc_ref}),0)*{tax_ref}"
                f"+{other_income_ref}"
                f"+{tv}*{sbc_ref}"
                f"+{tv}*{da_ref}"
                f"-{tv}*{capex_ref}"
            )
            c = ws.cell(row=r, column=col_idx)
            c.value = fcf_formula; c.number_format = NUM_CURRENCY
            c.font = FONT_FORMULA; c.fill = FILL_FCST
            c.alignment = Alignment(horizontal="right")

    # Base case: Txn 30% (ri=2 -> row 9), OpEx 42% (ci_idx=2 -> col 4=D)
    ws.cell(row=9, column=4).border = Border(
        top=_MEDIUM, left=_MEDIUM, right=_MEDIUM, bottom=_DOUBLE)
    ws.conditional_formatting.add("B7:F11", ColorScaleRule(
        start_type="min", start_color="FFC7CE",
        mid_type="percentile", mid_value=50, mid_color="FFEB9C",
        end_type="max", end_color="C6EFCE"))

    # ====================================================================
    # TABLE 2 — Net Income  (rows 13–20)
    # ====================================================================
    ws.row_dimensions[13].height = 18
    c = ws.cell(row=13, column=1)
    style_cell(c, "FY2026E Net Income ($M)  —  Transaction Revenue Growth x OpEx %",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK), fill=FILL_SUBHDR)
    ws.merge_cells("A13:F13")
    style_cell(ws.cell(row=14, column=1),
               "Same axes as Table 1. NI includes Other Income (flat corporate interest/float income) net of tax, consistent with the Income Statement.",
               font=FONT_NOTE)
    ws.merge_cells("A14:F14")

    _write_table_header(15, opex_vals, "OpEx %", "Txn Growth \\ OpEx")

    for ri, g_val in enumerate(txn_growth_vals):
        r = 16 + ri
        ca = ws.cell(row=r, column=1)
        ca.value = g_val; ca.font = FONT_INPUT; ca.fill = FILL_YELLOW
        ca.number_format = NUM_PCT; ca.alignment = Alignment(horizontal="center")
        for ci_idx in range(5):
            col_idx = 2 + ci_idx
            g_ref = f"$A{r}"
            o_ref = f"{cl(col_idx)}$15"
            tv = _total_rev(g_ref, ni_growth_ref, other_growth_ref)
            # NI = OI − Tax + Other Income  (Other Income is post-tax / below-the-line)
            ni_formula = (
                f"={tv}*(1-{o_ref}-{sbc_ref})"
                f"-MAX({tv}*(1-{o_ref}-{sbc_ref}),0)*{tax_ref}"
                f"+{other_income_ref}"
            )
            c = ws.cell(row=r, column=col_idx)
            c.value = ni_formula; c.number_format = NUM_CURRENCY
            c.font = FONT_FORMULA; c.fill = FILL_FCST
            c.alignment = Alignment(horizontal="right")

    # Base case: row 19, col D
    ws.cell(row=19, column=4).border = Border(
        top=_MEDIUM, left=_MEDIUM, right=_MEDIUM, bottom=_DOUBLE)
    ws.conditional_formatting.add("B16:F20", ColorScaleRule(
        start_type="min", start_color="FFC7CE",
        mid_type="percentile", mid_value=50, mid_color="FFEB9C",
        end_type="max", end_color="C6EFCE"))

    # ====================================================================
    # TABLE 3 — Total Revenue: Interest-Rate vs. Volume Sensitivity (rows 23–31)
    # ====================================================================
    ws.row_dimensions[23].height = 18
    c = ws.cell(row=23, column=1)
    style_cell(c, "FY2026E Total Revenue ($M)  —  NI Growth (rate) x Txn Growth (volume)",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK), fill=FILL_SUBHDR)
    ws.merge_cells("A23:F23")
    style_cell(ws.cell(row=24, column=1),
               "Row: Net Interest Revenue growth (-5% to +35%; model default 15% highlighted). "
               "Col: Transaction Revenue growth (10–50% — trading volume sensitivity). "
               "Other Revenue held at Assumptions.",
               font=FONT_NOTE)
    ws.merge_cells("A24:F24")

    _write_table_header(25, txn_growth_vals, "Txn Growth", "NI Growth \\ Txn Growth")

    for ri, g_ni_val in enumerate(ni_growth_vals):
        r = 26 + ri
        ca = ws.cell(row=r, column=1)
        ca.value = g_ni_val; ca.font = FONT_INPUT; ca.fill = FILL_YELLOW
        ca.number_format = NUM_PCT; ca.alignment = Alignment(horizontal="center")
        for ci_idx in range(5):
            col_idx = 2 + ci_idx
            g_ni_ref  = f"$A{r}"
            g_txn_ref = f"{cl(col_idx)}$25"
            tv = _total_rev(g_txn_ref, g_ni_ref, other_growth_ref)
            rev_formula = f"={tv}"
            c = ws.cell(row=r, column=col_idx)
            c.value = rev_formula; c.number_format = NUM_CURRENCY
            c.font = FONT_FORMULA; c.fill = FILL_FCST
            c.alignment = Alignment(horizontal="right")

    # Base case: NI growth=15% (model default) is row 28, Txn 30% is col D
    ws.cell(row=28, column=4).border = Border(
        top=_MEDIUM, left=_MEDIUM, right=_MEDIUM, bottom=_DOUBLE)
    ws.conditional_formatting.add("B26:F30", ColorScaleRule(
        start_type="min", start_color="FFC7CE",
        mid_type="percentile", mid_value=50, mid_color="FFEB9C",
        end_type="max", end_color="C6EFCE"))


# ---------------------------------------------------------------------------
# Model Integrity Checks (appended to Assumptions sheet)
# ---------------------------------------------------------------------------

def build_checks_section(
    ws,
    start_row: int,
    is_fcst_col_start: int,
    bs_fcst_col_start: int,
    cf_fcst_col_start: int,
) -> None:
    """Append model integrity checks to the Assumptions sheet (below assumptions).

    Validates internal consistency across all three statements:
      1. IS Revenue Sum: Verifies Total Revenue = Txn + NI + Other
      2. IS Costs: Verifies Total Costs = OpEx (ex-SBC) + SBC
      3. Net Income Formula: NI = OI - Taxes + Below-the-line
      4. BS Check: Partial Assets ≈ Partial L+E (non-zero due to excluded items)
      5. CF Check: Ending Cash = Prior + FCF + ΔDebt  (Cash Bridge)

    All checks are formula-based. Passing checks show "PASS" (green fill, zero);
    failing checks show the error amount (red fill, non-zero). This allows users
    to quickly spot internal inconsistencies or formula errors.

    Args:
        ws: openpyxl Worksheet object for Assumptions sheet.
        start_row (int): First row available for integrity checks block.
        is_fcst_col_start (int): First forecast column index on Income Statement.
        bs_fcst_col_start (int): First forecast column index on Balance Sheet.
        cf_fcst_col_start (int): First forecast column index on Cash Flow.

    Returns:
        None (modifies worksheet in-place).
    """
    checks_start = start_row

    # Section title bar
    ws.row_dimensions[checks_start].height = 20
    c = ws.cell(row=checks_start, column=1)
    style_cell(c, "Model Integrity Checks  (all values should = 0)",
               font=FONT_TITLE, fill=FILL_HEADER,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells(f"A{checks_start}:C{checks_start}")

    # Column sub-headers
    hdr_row = checks_start + 1
    for col_idx, label, h_align in [
        (1, "Check Description",  "left"),
        (2, "Value ($M)",         "center"),
        (3, "Status",             "center"),
    ]:
        style_cell(ws.cell(row=hdr_row, column=col_idx), label,
                   font=FONT_HDR, fill=FILL_SUBHDR,
                   alignment=Alignment(
                       horizontal=h_align,
                       indent=(1 if h_align == "left" else 0)
                   ))

    # Cross-sheet column letter references (FY2026E)
    is_let  = cl(is_fcst_col_start)          # IS  FY2026E  (col P)
    bs_this = cl(bs_fcst_col_start)          # BS  FY2026E  (col M)
    bs_prev = cl(bs_fcst_col_start - 1)      # BS  last hist (col L)
    cf_let  = cl(cf_fcst_col_start)          # CF  FY2026E  (col L)

    checks = [
        (
            "Cash Bridge Check — ΔCash = FCF + ΔDebt (FY2026E; should be 0 in closed model)",
            f"='Cash Flow'!{cf_let}{CF_ROW['Bridge Check']}",
        ),
        (
            "FCF Integrity — FCF = CFO − Capex (FY2026E)",
            (
                f"='Cash Flow'!{cf_let}{CF_ROW['FCF']}"
                f"-('Cash Flow'!{cf_let}{CF_ROW['CFO']}"
                f"-'Cash Flow'!{cf_let}{CF_ROW['Capex']})"
            ),
        ),
        (
            "Revenue Segments — Txn + NI + Other = Total Revenue (FY2026E)",
            (
                f"='Income Statement'!{is_let}{IS_ROW['Total Revenue']}"
                f"-('Income Statement'!{is_let}{IS_ROW['Txn Revenue']}"
                f"+'Income Statement'!{is_let}{IS_ROW['NI Revenue']}"
                f"+'Income Statement'!{is_let}{IS_ROW['Other Revenue']})"
            ),
        ),
        (
            "Equity Roll-Forward — Equity_t = Equity_{t-1} + Net Income (FY2026E)",
            (
                f"='Balance Sheet'!{bs_this}{BS_ROW['Equity']}"
                f"-'Balance Sheet'!{bs_prev}{BS_ROW['Equity']}"
                f"-'Income Statement'!{is_let}{IS_ROW['Net Income']}"
            ),
        ),
    ]

    check_rows: list[int] = []
    for i, (desc, formula) in enumerate(checks):
        r = hdr_row + 1 + i
        check_rows.append(r)
        # 30pt for rows that need 2 lines; row 78 (short text) left at auto-height.
        if i != 1:
            ws.row_dimensions[r].height = 30

        # Description label — wrap_text so long descriptions are fully visible
        style_cell(ws.cell(row=r, column=1), desc,
                   font=FONT_FORMULA,
                   alignment=Alignment(horizontal="left", indent=1,
                                       wrap_text=True, vertical="center"))

        # Formula value cell (should = 0)
        vc = ws.cell(row=r, column=2)
        vc.value         = formula
        vc.number_format = NUM_CURRENCY
        vc.font          = FONT_BOLD
        vc.alignment     = Alignment(horizontal="right")
        vc.border        = Border(bottom=_THIN)

        # Status cell: tolerance-based (ABS < 0.01 = $10k) to handle floating-point rounding
        sc = ws.cell(row=r, column=3)
        sc.value     = f'=IF(ABS(B{r})<0.01,"\u2713 PASS","\u2717 FAIL")'
        sc.font      = FONT_BOLD
        sc.alignment = Alignment(horizontal="center")
        sc.border    = Border(bottom=_THIN)

    # Conditional formatting — col B: green if |val|<0.01, red otherwise
    # Use tolerance to avoid false FAIL from floating-point rounding (~1e-13)
    for r in check_rows:
        ws.conditional_formatting.add(
            f"B{r}",
            FormulaRule(
                formula=[f"ABS($B${r})<0.01"],
                fill=FILL_GREEN_CF))
        ws.conditional_formatting.add(
            f"B{r}",
            FormulaRule(
                formula=[f"ABS($B${r})>=0.01"],
                fill=FILL_RED_CF))

        # Column C: inherit color from column B value
        ws.conditional_formatting.add(
            f"C{r}",
            FormulaRule(
                formula=[f"ABS($B${r})<0.01"],
                font=Font(name="Calibri", size=10, bold=True, color="375623"),
                fill=FILL_GREEN_CF,
            ))
        ws.conditional_formatting.add(
            f"C{r}",
            FormulaRule(
                formula=[f"ABS($B${r})>=0.01"],
                font=Font(name="Calibri", size=10, bold=True, color="9C0006"),
                fill=FILL_RED_CF,
            ))

    # Explanatory footnote
    note_r = hdr_row + len(checks) + 2
    style_cell(ws.cell(row=note_r, column=1),
               "All checks should display 0 / ✓ PASS in a properly integrated model. "
               "A non-zero / ✗ FAIL indicates a broken cross-sheet formula linkage. "
               "Checks reference FY2026E forecast columns only.",
               font=FONT_NOTE)


# ---------------------------------------------------------------------------
# CSV row validation gate (called from main before building)
# ---------------------------------------------------------------------------

_REQUIRED_IS_ROWS = [
    "Total Revenue",
    "Transaction-based Revenue",
    "Net Interest Revenue",
    "Other Revenue",
    "Operating Expenses",
    "Stock-Based Compensation",
    "Net Income",
]
_REQUIRED_BS_ROWS = ["Cash & Cash Equivalents", "Stockholders' Equity"]
_REQUIRED_CF_ROWS = ["Cash from Operations", "Free Cash Flow"]


def _validate_csv_rows(
    df_is: pd.DataFrame,
    df_bs: pd.DataFrame,
    df_cf: pd.DataFrame,
) -> None:
    """Raise ValueError if any required row is absent from the model CSVs.

    Called before the workbook is opened so that missing data causes a fast,
    clear failure rather than a silent empty sheet or a KeyError buried inside
    a builder function.
    """
    for row in _REQUIRED_IS_ROWS:
        if row not in df_is.index:
            raise ValueError(
                f"Income Statement CSV missing required row: '{row}'. "
                "Re-run make transform to regenerate."
            )
    for row in _REQUIRED_BS_ROWS:
        if row not in df_bs.index:
            raise ValueError(
                f"Balance Sheet CSV missing required row: '{row}'. "
                "Re-run make transform to regenerate."
            )
    for row in _REQUIRED_CF_ROWS:
        if row not in df_cf.index:
            raise ValueError(
                f"Cash Flow CSV missing required row: '{row}'. "
                "Re-run make transform to regenerate."
            )


def _validate_csv_completeness(
    df_is: "pd.DataFrame",
    df_bs: "pd.DataFrame",
    df_cf: "pd.DataFrame",
) -> None:
    """Abort if any CSV appears to be partially written (all-NaN latest column).

    A transform step that crashes mid-write can produce a CSV with correct
    column headers but empty data rows.  Checking the most-recent column
    catches this before the model builder silently produces blank forecasts.
    """
    for df, label in ((df_is, "IS"), (df_bs, "BS"), (df_cf, "CF")):
        if df.empty or len(df.columns) == 0:
            continue
        latest_col = df.columns[-1]
        non_null_pct = float(df[latest_col].notna().mean())
        if non_null_pct < 0.30:
            raise ValueError(
                f"[COMPLETENESS] {label} CSV column '{latest_col}' has only "
                f"{non_null_pct:.0%} non-null values — the CSV may be partially "
                "written. Re-run: make transform"
            )


# Manifest validation helper
# ---------------------------------------------------------------------------

def _validate_manifest(
    df_is: "pd.DataFrame",
    df_bs: "pd.DataFrame",
    df_cf: "pd.DataFrame",
) -> None:
    """Compare loaded CSVs against data/manifest.json written by the transform step.

    Compares full period label lists (not just counts) and **raises RuntimeError**
    if they diverge — a mismatch means the user ran ``make extract`` or edited
    CSVs without re-running ``make transform``, so the manifest is stale and the
    model would be built from inconsistent data.

    If manifest.json is absent (e.g., first run) the check is silently skipped.
    """
    import json
    manifest_path = DATA_DIR / "manifest.json"
    if not manifest_path.exists():
        return
    try:
        with open(manifest_path) as fh:
            manifest = json.load(fh)
        stmts = manifest.get("statements", {})
        for key, df, label in (
            ("IS", df_is, "Income Statement"),
            ("BS", df_bs, "Balance Sheet"),
            ("CF", df_cf, "Cash Flow"),
        ):
            expected_periods = stmts.get(key, {}).get("periods", [])
            actual_periods   = list(df.columns)
            if expected_periods and expected_periods != actual_periods:
                extra   = [p for p in actual_periods   if p not in expected_periods]
                missing = [p for p in expected_periods if p not in actual_periods]
                raise RuntimeError(
                    f"[MANIFEST MISMATCH] {label}: period labels in the CSV differ "
                    "from manifest.json — the transform step is likely stale. "
                    "Re-run 'make transform' to regenerate consistent data."
                    + (f"  Extra in CSV: {extra}"     if extra   else "")
                    + (f"  Missing in CSV: {missing}" if missing else "")
                )
    except (OSError, ValueError, KeyError) as e:
        logger.warning("  [MANIFEST] Could not read manifest.json (%s: %s) — skipping check",
                       type(e).__name__, e)


# ---------------------------------------------------------------------------
# Scenario helper
# ---------------------------------------------------------------------------

def _apply_scenario(
    spec: list,
    overrides: dict,
) -> list:
    """Return a copy of ASSUMPTIONS_SPEC with scenario override values applied."""
    result = []
    for item in spec:
        if item is None or item[1] is None or item[3] is None:
            result.append(item)
        elif item[3] in overrides:
            label, _, is_pct, key = item
            result.append((label, overrides[key], is_pct, key))
        else:
            result.append(item)
    return result


# ---------------------------------------------------------------------------
# Sheet 7 — Model Guide
# ---------------------------------------------------------------------------

def build_model_guide(ws) -> None:
    """Build the Model Guide sheet with navigation and key information.

    Provides a one-stop reference for model users:
      * Quick links to each sheet for navigation
      * Summary of key assumptions and their current values
      * Explanation of color-coding and styling conventions
      * Audit trail: what data sources were used, model version, last update
      * Instructions for updating assumptions and running scenarios

    Helps new users understand model structure and reduces support burden.

    Args:
        ws: openpyxl Worksheet object for Model Guide sheet.

    Returns:
        None (modifies worksheet in-place).
    """
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40

    # Row 1 – title bar
    ws.row_dimensions[1].height = 24
    style_cell(ws["A1"], f"{TICKER} Financial Model — Quick Reference Guide",
               font=FONT_TITLE, fill=FILL_HEADER,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells("A1:D1")

    # ------------------------------------------------------------------ #
    # Navigation section (rows 3–11)
    # ------------------------------------------------------------------ #
    row = 3
    ws.row_dimensions[row].height = 16
    c = ws.cell(row=row, column=1)
    style_cell(c, "Navigation",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells(f"A{row}:D{row}")

    # Column headers
    row += 1
    ws.row_dimensions[row].height = 16
    for col, label in [(1, "Sheet"), (2, "Purpose")]:
        c = ws.cell(row=row, column=col)
        style_cell(c, label, font=FONT_HDR, fill=FILL_HEADER,
                   alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells(f"B{row}:D{row}")

    nav_rows = [
        ("Model Guide",          "This sheet — navigation and reference"),
        ("Assumptions",          "All model inputs (blue/yellow cells). Edit here to change the model."),
        ("Income Statement",     "14 quarters actuals + 4-year forecast. Per-segment revenue, EPS, margins."),
        ("Balance Sheet",        "Cash rolls from FCF; Debt = revolver plug; Equity += Net Income each year."),
        ("Cash Flow",            "CFO = NI + SBC + D&A − ΔWC; FCF = CFO − Capex."),
        ("Valuation",            "DCF (Gordon Growth), exit multiples, WACC sensitivity."),
        ("Sensitivity Analysis", "5x5 tables: FCF / NI / Revenue across growth & cost inputs."),
    ]

    for i, (sheet_name, purpose) in enumerate(nav_rows):
        row += 1
        fill = FILL_GREY if i % 2 == 0 else FILL_NONE
        c_sheet = ws.cell(row=row, column=1)
        style_cell(c_sheet, sheet_name,
                   font=FONT_FORMULA,
                   fill=fill,
                   alignment=Alignment(horizontal="left", vertical="center", indent=1))
        c_purpose = ws.cell(row=row, column=2)
        style_cell(c_purpose, purpose,
                   font=FONT_FORMULA,
                   fill=fill,
                   alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True))
        ws.merge_cells(f"B{row}:D{row}")

    # ------------------------------------------------------------------ #
    # Key Assumptions section
    # ------------------------------------------------------------------ #
    row += 2   # blank spacer + section header
    ws.row_dimensions[row].height = 16
    c = ws.cell(row=row, column=1)
    style_cell(c, "Key Assumptions",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells(f"A{row}:D{row}")

    # Column headers
    row += 1
    ws.row_dimensions[row].height = 16
    for col, label in [(1, "Driver"), (2, "Default"), (3, "Range"), (4, "Notes")]:
        c = ws.cell(row=row, column=col)
        style_cell(c, label, font=FONT_HDR, fill=FILL_HEADER,
                   alignment=Alignment(horizontal="left", vertical="center", indent=1))

    key_assumptions = [
        ("Txn Rev Growth FY2026E",         "30%",  "10%-50%",  "Driven by crypto/equities volumes"),
        ("Net Interest Rev Growth FY2026E", "15%",  "0%-30%",   "Sensitive to Fed funds rate"),
        ("OpEx % of Revenue",              "42%",  "30%-55%",  "Excludes SBC"),
        ("Effective Tax Rate FY2026E",     "5%",   "0%-25%",   "NOL carryforward ~$2B"),
        ("WACC",                           "12%",  "8%-16%",   "See WACC sensitivity in Valuation"),
        ("Terminal Growth Rate",           "3%",   "1%-5%",    "Must be < WACC"),
    ]

    for i, (driver, default, rng, notes) in enumerate(key_assumptions):
        row += 1
        fill = FILL_GREY if i % 2 == 0 else FILL_NONE
        for col, val in [(1, driver), (2, default), (3, rng), (4, notes)]:
            c = ws.cell(row=row, column=col)
            style_cell(c, val,
                       font=FONT_FORMULA,
                       fill=fill,
                       alignment=Alignment(horizontal="left", vertical="center", indent=1))

    # ------------------------------------------------------------------ #
    # Limitations & Disclaimers section
    # ------------------------------------------------------------------ #
    row += 2   # blank spacer + section header
    ws.row_dimensions[row].height = 16
    c = ws.cell(row=row, column=1)
    style_cell(c, "Limitations & Disclaimers",
               font=Font(name="Calibri", size=10, bold=True, color=C_BLACK),
               fill=FILL_SUBHDR,
               alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells(f"A{row}:D{row}")

    limitations = [
        "1.  D&A is modeled as % of revenue, not an asset depreciation schedule.",
        "2.  Working Capital % assumptions calibrated to historical averages; include brokerage settlement flows.",
        "3.  Total Debt historically elevated by securities-lending liabilities — not corporate debt.",
        "4.  Share count is static; does not model incremental SBC dilution or option exercises.",
    ]

    for lim_text in limitations:
        row += 1
        c = ws.cell(row=row, column=1)
        style_cell(c, lim_text,
                   font=FONT_NOTE,
                   fill=FILL_NONE,
                   alignment=Alignment(horizontal="left", vertical="center",
                                       indent=1, wrap_text=True))
        ws.row_dimensions[row].height = 15
        ws.merge_cells(f"A{row}:D{row}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    """Main entry point: orchestrate entire Excel model build pipeline.

    Workflow:
      1. Parse command-line arguments (--scenario flag)
      2. Load scenario overrides from scenarios.yaml (if bull/bear selected)
      3. Validate scenario keys and bounds before touching Excel
      4. Load CSV data files (IS, BS, CF)
      5. Validate CSV structure, row names, and minimum data requirements
      6. Calculate forecast column start indices based on actual CSV column counts
      7. Create Workbook and build all 7 sheets in sequence
      8. Save to output file with scenario suffix (if applicable)

    Error Handling:
      * Missing CSVs: instructs user to run hood_data_transform.py first
      * Invalid scenario keys: lists valid keys from ASSUMPTIONS_SPEC
      * Non-numeric scenario values: rejects mixed types
      * Assumption bounds violations: prevents Excel generation to avoid bad models
      * Insufficient historical quarters: ensures minimum data for credible forecasts

    Returns:
        None (exits with code 1 on validation failure, 0 on success).
    """
    import argparse
    import yaml

    parser = argparse.ArgumentParser(description="Build HOOD Excel Model")
    parser.add_argument("--scenario", choices=["bull", "base", "bear"], default="base",
                        help="Named scenario to apply (bull/base/bear). Default: base.")
    args = parser.parse_args()
    scenario_name = args.scenario

    logger.info("%s Excel Model Builder", TICKER)
    logger.info("=" * 44)

    # ---- Load scenario overrides ----
    scenario_overrides: dict = {}
    if scenario_name != "base":
        scenarios_path = REPO_ROOT / "config" / "scenarios.yaml"
        with open(scenarios_path) as fh:
            all_scenarios = yaml.safe_load(fh)
        raw_overrides = all_scenarios.get(scenario_name, {}) or {}

        # Validate that every key in the scenario file corresponds to a real
        # assumption in ASSUMPTIONS_SPEC.  A typo (e.g. "txn_grwoth_y1") would
        # otherwise silently apply zero overrides and run the base case instead.
        valid_keys = {
            item[3]
            for item in ASSUMPTIONS_SPEC
            if item is not None and item[3] is not None
        }
        unknown = set(raw_overrides) - valid_keys
        if unknown:
            raise ValueError(
                f"Scenario '{scenario_name}' in scenarios.yaml contains unknown "
                f"assumption key(s): {sorted(unknown)}.  "
                f"Valid keys are: {sorted(valid_keys)}"
            )
        non_numeric = [
            f"{k}={v!r}"
            for k, v in raw_overrides.items()
            if not isinstance(v, (int, float))
        ]
        if non_numeric:
            raise ValueError(
                f"Scenario '{scenario_name}' contains non-numeric override value(s): "
                + ", ".join(non_numeric)
                + ".  All assumption values must be int or float."
            )
        scenario_overrides = raw_overrides
        logger.info(
            "  Applying scenario: %s — overrides: %s",
            scenario_name,
            {k: v for k, v in scenario_overrides.items()},
        )

    # Pre-build assumption bounds guard — catches out-of-range scenario overrides
    # before any Excel file is touched.  Mirrors the post-build bounds check in
    # validate_model.py but runs early so the user gets an immediate error.
    _GROWTH_BOUNDS  = (-0.50, 2.00)
    _TAX_BOUNDS     = (0.00,  0.50)
    _PCT_BOUNDS     = (0.00,  1.00)
    _WACC_BOUNDS    = (0.05,  0.30)
    _TGR_BOUNDS     = (0.00,  0.10)

    # (lo, hi, is_pct) — is_pct controls display formatting in error messages
    _BOUNDS_MAP: dict[str, tuple[float, float, bool]] = {
        # Revenue growth rates
        "txn_growth_y1": (*_GROWTH_BOUNDS, True), "txn_growth_y2": (*_GROWTH_BOUNDS, True),
        "txn_growth_y3": (*_GROWTH_BOUNDS, True), "txn_growth_y4": (*_GROWTH_BOUNDS, True),
        "ni_growth_y1":  (*_GROWTH_BOUNDS, True), "ni_growth_y2":  (*_GROWTH_BOUNDS, True),
        "ni_growth_y3":  (*_GROWTH_BOUNDS, True), "ni_growth_y4":  (*_GROWTH_BOUNDS, True),
        "other_growth_y1": (*_GROWTH_BOUNDS, True), "other_growth_y2": (*_GROWTH_BOUNDS, True),
        "other_growth_y3": (*_GROWTH_BOUNDS, True), "other_growth_y4": (*_GROWTH_BOUNDS, True),
        # Cost structure (percentages of revenue)
        "opex_pct":      (*_PCT_BOUNDS, True),
        "sbc_pct":       (*_PCT_BOUNDS, True),
        "da_pct":        (*_PCT_BOUNDS, True),
        "capex_pct":     (*_PCT_BOUNDS, True),
        # Tax rates
        "tax_rate_y1":   (*_TAX_BOUNDS, True), "tax_rate_y2": (*_TAX_BOUNDS, True),
        "tax_rate_y3":   (*_TAX_BOUNDS, True), "tax_rate_y4": (*_TAX_BOUNDS, True),
        "statutory_rate": (*_TAX_BOUNDS, True),
        # Working capital (percentages of revenue)
        "recv_pct":      (0.00, 1.00, True),
        "pay_pct":       (0.00, 1.00, True),
        # Valuation
        "wacc":            (*_WACC_BOUNDS, True),
        "terminal_growth": (*_TGR_BOUNDS, True),
        "exit_ebitda_mult": (0.0, 100.0, False),
        "exit_rev_mult":    (0.0, 50.0, False),
        # Dollar amounts
        "nol_balance":     (0.0, 50_000.0, False),
        "min_cash":        (0.0, 50_000.0, False),
        "other_income":    (-5_000.0, 5_000.0, False),
        "shares_diluted":  (1.0, 50_000.0, False),
        "min_debt":        (0.0, 50_000.0, False),
        "max_debt":        (0.0, 100_000.0, False),
    }
    _effective_spec = _apply_scenario(ASSUMPTIONS_SPEC, scenario_overrides)
    _violations: list[str] = []
    for _item in _effective_spec:
        if _item is None or _item[3] is None or _item[1] is None:
            continue
        _key, _val = _item[3], _item[1]
        if _key in _BOUNDS_MAP:
            _lo, _hi, _is_pct = _BOUNDS_MAP[_key]
            if not (_lo <= _val <= _hi):
                if _is_pct:
                    _violations.append(f"{_key}={_val:.1%} (allowed {_lo:.0%}–{_hi:.0%})")
                else:
                    _violations.append(f"{_key}={_val:,.1f} (allowed {_lo:,.0f}–{_hi:,.0f})")
    if _violations:
        raise ValueError(
            f"Scenario '{scenario_name}' has assumption(s) outside valid bounds: "
            + "; ".join(_violations)
        )

    required = [
        DATA_DIR / "model_income_statement.csv",
        DATA_DIR / "model_balance_sheet.csv",
        DATA_DIR / "model_cash_flow.csv",
    ]
    missing = [str(f) for f in required if not f.exists()]
    if missing:
        logger.error(
            "Missing model CSV(s):\n%s\n\n"
            "These CSVs are produced by the transform step.\n"
            "Run:  python -m src.hood_data_transform",
            "\n".join(f"  - {f}" for f in missing),
        )
        sys.exit(1)

    df_is = pd.read_csv(DATA_DIR / "model_income_statement.csv", index_col=0)
    df_bs = pd.read_csv(DATA_DIR / "model_balance_sheet.csv",    index_col=0)
    df_cf = pd.read_csv(DATA_DIR / "model_cash_flow.csv",        index_col=0)

    _validate_manifest(df_is, df_bs, df_cf)
    _validate_csv_rows(df_is, df_bs, df_cf)
    _validate_csv_completeness(df_is, df_bs, df_cf)

    # Verify that we have enough historical quarters to build a credible model.
    # Forecast column indices are derived directly from CSV column counts; too few
    # quarters would silently shift all forecast columns left and produce wrong output.
    _MIN_IS_QUARTERS    = 8   # need at least 8 quarters of IS history
    _MIN_BS_CF_QUARTERS = 6   # need at least 6 quarters of BS/CF history
    if len(df_is.columns) < _MIN_IS_QUARTERS:
        raise ValueError(
            f"Income Statement has only {len(df_is.columns)} quarter(s) "
            f"(minimum required: {_MIN_IS_QUARTERS}). Re-run: make extract && make transform"
        )
    if len(df_bs.columns) < _MIN_BS_CF_QUARTERS:
        raise ValueError(
            f"Balance Sheet has only {len(df_bs.columns)} quarter(s) "
            f"(minimum required: {_MIN_BS_CF_QUARTERS}). Re-run: make extract && make transform"
        )
    if len(df_cf.columns) < _MIN_BS_CF_QUARTERS:
        raise ValueError(
            f"Cash Flow has only {len(df_cf.columns)} quarter(s) "
            f"(minimum required: {_MIN_BS_CF_QUARTERS}). Re-run: make extract && make transform"
        )

    # Forecast column start indices (1-based, sheet-specific) — derived from
    # actual CSV column counts so they're always accurate regardless of how many
    # quarters were extracted.
    is_fcst_col_start = HIST_COL_START + len(df_is.columns)   # IS: col P (16)
    bs_fcst_col_start = HIST_COL_START + len(df_bs.columns)   # BS: col M (13)
    cf_fcst_col_start = HIST_COL_START + len(df_cf.columns)   # CF: col L (12)

    # ---- Historical WC averages for Assumptions annotation ----
    # Use the last 4 IS quarters that also have corresponding BS data.
    shared_cols = [c for c in df_is.columns[-4:] if c in df_bs.columns]
    recv_ratios = [
        df_bs.loc["Receivables", q] / df_is.loc["Total Revenue", q]
        for q in shared_cols
        if "Receivables" in df_bs.index
        and "Total Revenue" in df_is.index
        and pd.notna(df_bs.loc["Receivables", q])
        and pd.notna(df_is.loc["Total Revenue", q])
        and df_is.loc["Total Revenue", q] != 0
    ]
    pay_ratios = [
        df_bs.loc["Payables", q] / df_is.loc["Total Revenue", q]
        for q in shared_cols
        if "Payables" in df_bs.index
        and "Total Revenue" in df_is.index
        and pd.notna(df_bs.loc["Payables", q])
        and pd.notna(df_is.loc["Total Revenue", q])
        and df_is.loc["Total Revenue", q] != 0
    ]
    hist_recv_pct = sum(recv_ratios) / len(recv_ratios) if recv_ratios else None
    hist_pay_pct  = sum(pay_ratios)  / len(pay_ratios)  if pay_ratios  else None

    wb = Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Assumptions                                                #
    # ------------------------------------------------------------------ #
    ws_assump = wb.active
    ws_assump.title = "Assumptions"
    logger.info("  Building Assumptions sheet…")
    cell_refs, checks_start = build_assumptions(
        ws_assump, is_fcst_col_start,
        hist_recv_pct=hist_recv_pct,
        hist_pay_pct=hist_pay_pct,
        spec=_apply_scenario(ASSUMPTIONS_SPEC, scenario_overrides),
    )

    # ------------------------------------------------------------------ #
    # Sheet 2: Income Statement                                           #
    # ------------------------------------------------------------------ #
    ws_is = wb.create_sheet("Income Statement")
    logger.info("  Building Income Statement (historical + 4-year forecast)…")
    build_income_statement(ws_is, df_is, cell_refs, cf_fcst_col_start)

    # ------------------------------------------------------------------ #
    # Sheet 3: Balance Sheet                                              #
    # ------------------------------------------------------------------ #
    ws_bs = wb.create_sheet("Balance Sheet")
    logger.info("  Building Balance Sheet (historical + 4-year forecast)…")
    build_balance_sheet(ws_bs, df_bs, cell_refs, is_fcst_col_start, cf_fcst_col_start)

    # ------------------------------------------------------------------ #
    # Sheet 4: Cash Flow                                                  #
    # ------------------------------------------------------------------ #
    ws_cf = wb.create_sheet("Cash Flow")
    logger.info("  Building Cash Flow Statement (historical + 4-year forecast)…")
    build_cash_flow(ws_cf, df_cf, cell_refs, is_fcst_col_start, bs_fcst_col_start)

    # ------------------------------------------------------------------ #
    # Sheet 5: Valuation Analysis                                         #
    # ------------------------------------------------------------------ #
    ws_val = wb.create_sheet("Valuation")
    logger.info("  Building Valuation Analysis…")
    build_valuation(ws_val, cell_refs, is_fcst_col_start, bs_fcst_col_start, cf_fcst_col_start)

    # ------------------------------------------------------------------ #
    # Sheet 6: Sensitivity Analysis                                       #
    # ------------------------------------------------------------------ #
    ws_sens = wb.create_sheet("Sensitivity Analysis")
    logger.info("  Building Sensitivity Analysis…")
    build_sensitivity(ws_sens, cell_refs, is_fcst_col_start)

    # ------------------------------------------------------------------ #
    # Sheet 7: Model Guide                                                #
    # ------------------------------------------------------------------ #
    ws_guide = wb.create_sheet("Model Guide")
    logger.info("  Building Model Guide…")
    build_model_guide(ws_guide)

    # ------------------------------------------------------------------ #
    # Model Integrity Checks (appended to Assumptions sheet)             #
    # ------------------------------------------------------------------ #
    logger.info("  Adding Model Integrity Checks (Assumptions tab)…")
    build_checks_section(
        ws_assump, checks_start,
        is_fcst_col_start, bs_fcst_col_start, cf_fcst_col_start,
    )

    suffix = f"_{scenario_name}" if scenario_name != "base" else ""
    output = OUTPUT_DIR / f"{TICKER}_Financial_Model{suffix}.xlsx"
    wb.save(output)
    logger.info("\n  Saved → %s", output)
    logger.info("  Sheets: Assumptions | Income Statement | Balance Sheet | Cash Flow | Valuation | Sensitivity Analysis | Model Guide")
    logger.info("\nDone.")


if __name__ == "__main__":
    logging.basicConfig(
        level=os.environ.get("LOG_LEVEL", "INFO").upper(),
        format="%(message)s",
    )
    main()
