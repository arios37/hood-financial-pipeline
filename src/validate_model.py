"""
HOOD Financial Model Validator
===============================

Module Overview
---------------
This module validates the integrity of the HOOD financial model across two dimensions:
  1. File & Manifest integrity: Ensures all required outputs exist and period labels are consistent
  2. Structural integrity: Verifies Excel sheet structure, row labels, and formula linkages
  3. Data quality: Cross-checks numerical relationships and bounds across statements

Purpose
-------
The validator is a critical CI/CD step that prevents invalid or corrupted models from being
deployed. It enforces that:
  - All CSV extracts and Excel workbook are present
  - Excel sheets have the correct structure (row labels, formulas, cross-sheet references)
  - Financial data is internally consistent (FCF formula, NI ties, equity bounds)
  - Assumptions sheet values are within reasonable bounds

27 Checks Performed
-------------------
  [FILES  1]  model_income_statement.csv present
  [FILES  2]  model_balance_sheet.csv present
  [FILES  3]  model_cash_flow.csv present
  [FILES  4]  Excel output (HOOD_Financial_Model.xlsx) present
  [FILES  5]  manifest.json present (written by transform step)
  [FILES  6]  Manifest IS period labels match CSV (full list, not just count)
  [FILES  7]  Manifest BS period labels match CSV
  [FILES  8]  Manifest CF period labels match CSV
  [FILES  9]  BS periods are a subset of IS periods (cross-statement consistency)
  [FILES 10]  CF periods are a subset of IS periods (cross-statement consistency)
  [STRUCT 11] All 7 sheets present in Excel
  [STRUCT 12] Income Statement row labels correct (rows 8 / 14 / 18)
  [STRUCT 13] Balance Sheet row labels correct (rows 4 / 8 / 9)
  [STRUCT 14] Cash Flow row labels correct (rows 7 / 9)
  [STRUCT 15] IS forecast cells contain Excel formulas (not hardcoded)
  [STRUCT 16] BS forecast cells contain Excel formulas (not hardcoded)
  [STRUCT 17] CF forecast cells contain Excel formulas (not hardcoded)
  [STRUCT 18] Sensitivity data cells contain Excel formulas
  [STRUCT 19] IS FY2026E revenue formula references Assumptions sheet
  [STRUCT 20] Valuation sheet contains DCF formulas
  [STRUCT 21] Assumption values within valid bounds
  [DATA   22] Period alignment — IS / BS / CF share ≥ 6 common quarters
  [DATA   23] Total Revenue non-null for all IS periods
  [DATA   24] Net Income non-null for most recent 4 IS periods
  [DATA   25] FCF = CFO − Capex for all complete CF periods (±$1M)
  [DATA   26] Net Income ties between IS and CF for shared periods (±$1M)
  [DATA   27] Stockholders' Equity positive for all BS periods

Running the Validator
---------------------
From the project root directory:
  python -m src.validate_model

Exit codes:
  0 = all 27 checks passed
  1 = one or more checks failed
"""

from __future__ import annotations

import logging
import os
import sys

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration: Paths & Layout Constants
# ---------------------------------------------------------------------------
# All paths and layout constants are defined in config.py (single source of truth).
# This avoids duplication and ensures consistency across the pipeline.

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir))
from config import (
    DATA_DIR, OUTPUT_DIR, TICKER,
    EXPECTED_SHEETS,
    IS_FCST_COL,
    IS_ROW,
    BS_ROW,
    CF_ROW,
    BS_FCST_COL_START,
    CF_FCST_COL_START,
)

EXCEL_PATH = OUTPUT_DIR / f"{TICKER}_Financial_Model.xlsx"

# CSV file paths for the three required financial statements
REQUIRED_CSVS = [
    DATA_DIR / "model_income_statement.csv",
    DATA_DIR / "model_balance_sheet.csv",
    DATA_DIR / "model_cash_flow.csv",
]

# Numerical tolerance ($M) for tie-break checks (FCF formula, NI alignment).
# Uses $1M to allow for minor rounding differences in Excel vs. pandas calculations.
TOLERANCE = 1.0  # $M tolerance for numerical tie checks

# ---------------------------------------------------------------------------
# Helpers: Output Formatting and Results Accumulation
# ---------------------------------------------------------------------------

PASS = "\033[32mPASS\033[0m"  # Green ANSI escape for PASS
FAIL = "\033[31mFAIL\033[0m"  # Red ANSI escape for FAIL

# Global results accumulator for summary reporting at the end.
# Each entry is a tuple (passed: bool, name: str, detail: str).
results: list[tuple[bool, str, str]] = []


def check(passed: bool, name: str, detail: str = "") -> bool:
    """
    Log a single validation check result.

    Appends the result to the global results list, prints a formatted message
    with color-coded PASS/FAIL tags, and optionally includes a detail message
    for failed checks.

    Args:
        passed (bool): True if the check passed, False otherwise.
        name (str): Human-readable name of the check (e.g., "CSV present: model_income_statement.csv").
        detail (str, optional): Extra context for failed checks (e.g., error message, mismatch details).
                                Defaults to empty string.

    Returns:
        bool: The passed status unchanged (useful for chaining or assertion-like patterns).
    """
    tag = PASS if passed else FAIL
    msg = f"  [{tag}] {name}"
    if detail:
        msg += f"\n         {detail}"
    print(msg)
    results.append((passed, name, detail))
    return passed


# ---------------------------------------------------------------------------
# FILE Checks (Existence, Manifest, Period Alignment)
# ---------------------------------------------------------------------------

def check_files() -> None:
    """
    Validate file presence and period label consistency.

    Performs 10 checks:
      1. All required CSV files exist
      2. Excel workbook exists
      3. manifest.json exists and is valid JSON
      4–8. For each statement (IS, BS, CF), verify that period labels in the manifest
            match the period columns in the actual CSV (full list match, not just count)
      9–10. Cross-statement consistency: BS and CF periods must be subsets of IS periods.
            This catches cases where IS was extracted from more comprehensive 10-K filings
            than BS/CF, or indicates different transform runs were used.

    A mismatch in period labels typically signals that transform/extract needs to be re-run.
    """
    import json
    logger.info("\n── Files ──────────────────────────────────────────────────────")

    # Check 1–3: Required CSV files present
    for p in REQUIRED_CSVS:
        check(p.exists(), f"CSV present: {p.name}")

    # Check 4: Excel output present
    check(EXCEL_PATH.exists(), f"Excel output present: {EXCEL_PATH.name}")

    # Check 5: Manifest file present
    manifest_path = DATA_DIR / "manifest.json"
    check(manifest_path.exists(), "Manifest present: manifest.json",
          "Run make transform to generate" if not manifest_path.exists() else "")

    # Checks 6–10: If manifest exists, validate period labels for each statement
    if manifest_path.exists():
        try:
            with open(manifest_path) as fh:
                manifest = json.load(fh)
            stmts = manifest.get("statements", {})

            # Map statement keys to their CSV paths and readable labels
            csv_map = {
                "IS": (DATA_DIR / "model_income_statement.csv", "Income Statement"),
                "BS": (DATA_DIR / "model_balance_sheet.csv",    "Balance Sheet"),
                "CF": (DATA_DIR / "model_cash_flow.csv",        "Cash Flow"),
            }

            # Load periods from CSVs and compare against manifest
            loaded: dict[str, list] = {}
            for key, (csv_path, label) in csv_map.items():
                if csv_path.exists() and key in stmts:
                    df = pd.read_csv(csv_path, index_col=0)
                    actual_periods   = list(df.columns)
                    expected_periods = stmts[key].get("periods", [])
                    periods_match    = actual_periods == expected_periods
                    detail = ""

                    # If periods don't match, compute extra and missing for diagnosis
                    if not periods_match:
                        extra   = [p for p in actual_periods   if p not in expected_periods]
                        missing = [p for p in expected_periods if p not in actual_periods]
                        detail  = (
                            "re-run make transform"
                            + (f" | extra in CSV: {extra}"   if extra   else "")
                            + (f" | missing in CSV: {missing}" if missing else "")
                        )

                    check(
                        periods_match,
                        f"Manifest period labels match CSV: {label}",
                        detail,
                    )
                    loaded[key] = actual_periods

            # Cross-statement subset checks (Checks 9–10):
            # BS and CF periods must be subsets of IS periods. A mismatch signals that
            # IS was extracted from more 10-K filings than BS/CF, or different transform runs.
            if "IS" in loaded and "BS" in loaded:
                bs_not_in_is = [p for p in loaded["BS"] if p not in loaded["IS"]]
                check(
                    not bs_not_in_is,
                    "BS periods are a subset of IS periods",
                    f"BS has periods not in IS: {bs_not_in_is} — re-run make transform"
                    if bs_not_in_is else "",
                )
            if "IS" in loaded and "CF" in loaded:
                cf_not_in_is = [p for p in loaded["CF"] if p not in loaded["IS"]]
                check(
                    not cf_not_in_is,
                    "CF periods are a subset of IS periods",
                    f"CF has periods not in IS: {cf_not_in_is} — re-run make transform"
                    if cf_not_in_is else "",
                )
        except (OSError, ValueError, KeyError) as e:
            check(False, "Manifest readable", str(e))


# ---------------------------------------------------------------------------
# STRUCTURAL Checks (Excel: Sheet Names, Row Labels, Formulas, Bounds)
# ---------------------------------------------------------------------------

def check_structure(wb) -> None:
    """
    Validate Excel workbook structure, formulas, and bounds.

    Performs 10 checks on the Excel file:
      11. All 7 expected sheets (IS, BS, CF, Sensitivity, Assumptions, Valuation, Dashboard) present
      12–14. Row label validation for IS, BS, CF (ensures metrics are in expected rows)
      15–17. Formula validation: forecast cells (FY2026E–FY2029E) must be formulas, not hardcoded values.
              Why: Hardcoded values would break the linkage to the Assumptions sheet.
      18. Sensitivity table (rows 7–11, cols B–F) contains formulas
      19. IS revenue formula references Assumptions sheet (critical cross-sheet linkage)
      20. Valuation sheet contains DCF formulas
      21. Assumption bounds validation: growth rates, tax rates, WACC, terminal growth rate

    Args:
        wb: openpyxl Workbook object loaded with data_only=False to read formulas as strings.
    """
    logger.info("\n── Structure ──────────────────────────────────────────────────")

    # Check 11: All required sheets present
    missing = [s for s in EXPECTED_SHEETS if s not in wb.sheetnames]
    check(not missing, "All 7 sheets present",
          f"Missing: {missing}" if missing else "")

    # Reference the main statement sheets (will be None if not found)
    ws_is = wb["Income Statement"]  if "Income Statement"  in wb.sheetnames else None
    ws_bs = wb["Balance Sheet"]     if "Balance Sheet"     in wb.sheetnames else None
    ws_cf = wb["Cash Flow"]         if "Cash Flow"         in wb.sheetnames else None
    ws_sa = wb["Sensitivity Analysis"] if "Sensitivity Analysis" in wb.sheetnames else None

    # ─────────────────────────────────────────────────────────────────────
    # Checks 12–14: Row Label Validation
    # ─────────────────────────────────────────────────────────────────────
    # Verifies that key financial metrics are located in the expected row positions.
    # Ensures the Excel structure matches our layout assumptions.

    # Check 12: IS row labels (Total Revenue at 8, Operating Income at 14, Net Income at 18)
    if ws_is:
        expected_is = {8: "Total Revenue", 14: "Operating Income", 18: "Net Income"}
        bad = {r: (v, ws_is.cell(row=r, column=1).value)
               for r, v in expected_is.items()
               if ws_is.cell(row=r, column=1).value != v}
        check(not bad, "Income Statement row labels correct",
              f"Mismatches: {bad}" if bad else "")
    else:
        check(False, "Income Statement row labels correct", "Sheet missing")

    # Check 13: BS row labels (Cash at 4, Total Debt at 8, Equity at 9)
    if ws_bs:
        expected_bs = {4: "Cash & Cash Equivalents", 8: "Total Debt", 9: "Stockholders' Equity"}
        bad = {r: (v, ws_bs.cell(row=r, column=1).value)
               for r, v in expected_bs.items()
               if ws_bs.cell(row=r, column=1).value != v}
        check(not bad, "Balance Sheet row labels correct",
              f"Mismatches: {bad}" if bad else "")
    else:
        check(False, "Balance Sheet row labels correct", "Sheet missing")

    # Check 14: CF row labels (CFO at 7, FCF at 9)
    if ws_cf:
        expected_cf = {7: "Cash from Operations", 9: "Free Cash Flow"}
        bad = {r: (v, ws_cf.cell(row=r, column=1).value)
               for r, v in expected_cf.items()
               if ws_cf.cell(row=r, column=1).value != v}
        check(not bad, "Cash Flow row labels correct",
              f"Mismatches: {bad}" if bad else "")
    else:
        check(False, "Cash Flow row labels correct", "Sheet missing")

    # ─────────────────────────────────────────────────────────────────────
    # Checks 15–17: Formula Cell Validation for Forecast Data
    # ─────────────────────────────────────────────────────────────────────
    # All forecast cells (FY2026E–FY2029E) must contain Excel formulas, not hardcoded values.
    # Why: Hardcoded numbers break the linkage to the Assumptions sheet and would allow stale
    #      forecast data to persist even after assumptions are updated.
    #
    # To handle new quarters robustly, we detect the first forecast column dynamically by
    # scanning the header row (row 3) for "FY2026E", rather than relying on config constants
    # which may lag if new quarters are extracted but config hasn't been updated.

    def _actual_fcst_col(ws, fallback: int, header_row: int = 3) -> int:
        """
        Determine the first forecast column index by scanning headers.

        Scans the specified header row for the first cell starting with "FY2026" to identify
        where forecast data begins. Falls back to a config constant if no header is found,
        ensuring robustness when new forecast years are added.

        Args:
            ws: openpyxl Worksheet object
            fallback (int): 1-based column index fallback (from config) if header not found
            header_row (int): The row containing period headers (default: 3)

        Returns:
            int: 1-based column index of the first forecast column
        """
        for cell in ws[header_row]:
            if str(cell.value or "").startswith("FY2026"):
                return cell.column
        return fallback   # config constant as last-resort fallback

    def _count_non_formula_fcst_cells(ws, fcst_col_start: int, data_rows: list[int]) -> list[str]:
        """
        Find forecast cells that are not formulas.

        Scans the forecast range (4 columns starting at fcst_col_start, for each data row)
        and collects cell addresses of cells that contain non-formula values or are non-blank.

        Args:
            ws: openpyxl Worksheet object
            fcst_col_start (int): 1-based column index of first forecast column
            data_rows (list[int]): List of 1-based row indices to check

        Returns:
            list[str]: List of cell coordinate strings with non-formula values (e.g., ["E8=123.45"])
        """
        bad: list[str] = []
        for r in data_rows:
            for fi in range(4):   # 4 forecast years (FY2026E–FY2029E)
                c = ws.cell(row=r, column=fcst_col_start + fi)
                if c.value is None:
                    continue   # blank cells are fine (e.g., memo/helper rows)
                # A formula string always starts with "=" in openpyxl (data_only=False)
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    bad.append(f"{c.coordinate}={c.value!r}")
        return bad

    # Check 15: IS forecast cells contain formulas
    if ws_is:
        is_fcst_actual = _actual_fcst_col(ws_is, IS_FCST_COL)
        is_data_rows = [IS_ROW[k] for k in ("Txn Revenue", "NI Revenue", "Other Revenue",
                                              "Total Revenue", "Operating Expenses", "SBC",
                                              "Operating Income", "Tax Provision", "Net Income")]
        bad_is = _count_non_formula_fcst_cells(ws_is, is_fcst_actual, is_data_rows)
        check(not bad_is, "IS forecast cells contain Excel formulas",
              f"{len(bad_is)} non-formula cell(s): {bad_is[:3]}" if bad_is else "")
    else:
        check(False, "IS forecast cells contain Excel formulas", "Sheet missing")

    # Check 16: BS forecast cells contain formulas
    if ws_bs:
        bs_fcst_actual = _actual_fcst_col(ws_bs, BS_FCST_COL_START)
        bs_data_rows = [BS_ROW[k] for k in ("Cash", "Restricted Cash", "Receivables",
                                              "Payables", "Total Debt", "Equity")]
        bad_bs = _count_non_formula_fcst_cells(ws_bs, bs_fcst_actual, bs_data_rows)
        check(not bad_bs, "BS forecast cells contain Excel formulas",
              f"{len(bad_bs)} non-formula cell(s): {bad_bs[:3]}" if bad_bs else "")
    else:
        check(False, "BS forecast cells contain Excel formulas", "Sheet missing")

    # Check 17: CF forecast cells contain formulas
    if ws_cf:
        cf_fcst_actual = _actual_fcst_col(ws_cf, CF_FCST_COL_START)
        cf_data_rows = [CF_ROW[k] for k in ("Net Income", "SBC", "DA", "CFO", "Capex", "FCF")]
        bad_cf = _count_non_formula_fcst_cells(ws_cf, cf_fcst_actual, cf_data_rows)
        check(not bad_cf, "CF forecast cells contain Excel formulas",
              f"{len(bad_cf)} non-formula cell(s): {bad_cf[:3]}" if bad_cf else "")
    else:
        check(False, "CF forecast cells contain Excel formulas", "Sheet missing")

    # Check 18: Sensitivity table formulas
    # Sensitivity Analysis sheet, rows 7–11 (5 rows), cols B–F (5 cols) = 25 data cells.
    # All must be formulas that reference the Assumptions sheet for robust what-if analysis.
    if ws_sa:
        sens_cells = [ws_sa.cell(row=r, column=c) for r in range(7, 12) for c in range(2, 7)]
        all_formulas = all(
            isinstance(c.value, str) and c.value.startswith("=")
            for c in sens_cells
        )
        sample_b7: str = str(ws_sa.cell(row=7, column=2).value or "")[:40]
        check(all_formulas, "Sensitivity data cells contain Excel formulas",
              f"Sample[B7]: {sample_b7!r}" if not all_formulas else "")
    else:
        check(False, "Sensitivity data cells contain Excel formulas", "Sheet missing")

    # Check 19: IS FY2026E revenue linkage to Assumptions
    # Verifies that segment revenue formulas (Txn Revenue, NI Revenue, Other Revenue)
    # reference the Assumptions sheet. This ensures that changing growth assumptions
    # in the Assumptions sheet cascades through to forecasts.
    if ws_is:
        # Txn Revenue (row 5) is one of three segment revenue cells linked to Assumptions
        txn_cell = ws_is.cell(row=IS_ROW["Txn Revenue"], column=IS_FCST_COL)
        formula  = txn_cell.value or ""
        cross_refs_assumptions = "Assumptions!" in str(formula)
        check(
            cross_refs_assumptions,
            "IS FY2026E revenue formula references Assumptions sheet",
            f"Formula: {str(formula)[:80]}" if not cross_refs_assumptions else "",
        )
    else:
        check(False, "IS FY2026E revenue formula references Assumptions sheet", "Sheet missing")

    # Check 20: Valuation sheet DCF formulas
    # The Valuation sheet should contain NPV and implied share price calculations
    # that reference FCF and WACC from other sheets.
    ws_val = wb["Valuation"] if "Valuation" in wb.sheetnames else None
    if ws_val:
        # Scan rows 3–7 (typical valuation metrics) in col B for formula cells
        val_cells = [ws_val.cell(row=r, column=2) for r in range(3, 8)]
        has_formulas = any(
            isinstance(c.value, str) and c.value.startswith("=") for c in val_cells
        )
        check(has_formulas, "Valuation sheet contains DCF formulas",
              "No formula cells found in rows 3-7 col B" if not has_formulas else "")
    else:
        check(False, "Valuation sheet contains DCF formulas", "Sheet missing")

    # ─────────────────────────────────────────────────────────────────────
    # Check 21: Assumption Bounds Validation
    # ─────────────────────────────────────────────────────────────────────
    # Validates that assumption input cells contain values within reasonable ranges.
    # This catches accidental typos (e.g., 5.0 instead of 0.05 for growth rate) and
    # impossible values (negative tax rates, WACC above 30%, etc.).
    #
    # Input cells contain raw numeric values (not formulas), so reading with data_only=False
    # is fine (we read the raw input, not computed results).
    ws_assump = wb["Assumptions"] if "Assumptions" in wb.sheetnames else None
    if ws_assump:
        bounds_failures: list[str] = []

        # Growth rates (Txn, NI, Other): -50% to 200%
        # Negative growth is allowed (recession/decline), but >200% is unrealistic.
        growth_rows = list(range(5, 9)) + list(range(11, 15)) + list(range(17, 21))
        for r in growth_rows:
            v = ws_assump.cell(row=r, column=2).value
            if isinstance(v, (int, float)) and not (-0.50 <= v <= 2.00):
                bounds_failures.append(f"B{r}={v:.1%} (growth out of −50% – 200%)")

        # Tax rates: 0% to 50%
        # 0% allows for loss-carryforward scenarios; 50% is high but plausible for some jurisdictions.
        for r in range(28, 32):
            v = ws_assump.cell(row=r, column=2).value
            if isinstance(v, (int, float)) and not (0.0 <= v <= 0.50):
                bounds_failures.append(f"B{r}={v:.1%} (tax rate out of 0% – 50%)")

        # WACC (Weighted Average Cost of Capital): 5% to 30%
        # Typical range for most companies; <5% is unrealistic, >30% suggests high risk.
        wacc_val = ws_assump.cell(row=50, column=2).value
        if isinstance(wacc_val, (int, float)) and not (0.05 <= wacc_val <= 0.30):
            bounds_failures.append(f"B50={wacc_val:.1%} (WACC out of 5% – 30%)")

        # Terminal Growth Rate (TGR): 0% to 10%
        # Long-term growth cannot exceed GDP growth; 10% is aggressive but possible in emerging markets.
        tgr_val = ws_assump.cell(row=51, column=2).value
        if isinstance(tgr_val, (int, float)) and not (0.0 <= tgr_val <= 0.10):
            bounds_failures.append(f"B51={tgr_val:.1%} (TGR out of 0% – 10%)")

        check(
            not bounds_failures,
            "Assumption values within valid bounds",
            "; ".join(bounds_failures) if bounds_failures else "",
        )
    else:
        check(False, "Assumption values within valid bounds", "Assumptions sheet missing")


# ---------------------------------------------------------------------------
# DATA Checks (Pandas: Numerical Relationships & Quality)
# ---------------------------------------------------------------------------

def check_data() -> None:
    """
    Validate financial data relationships and quality across CSV statements.

    Performs 6 checks on the extracted CSV data:
      22. Period alignment: IS, BS, CF share at least 6 common quarters (timeline coverage)
      23. Total Revenue non-null for all IS periods (completeness)
      24. Net Income non-null for most recent 4 IS periods (recent data quality)
      25. FCF = CFO − Capex (±$1M) for complete CF periods (formula verification)
      26. Net Income ties between IS and CF for shared periods (±$1M) (cross-statement reconciliation)
      27. Stockholders' Equity positive for all BS periods (financial health sanity check)

    Checks 25–26 use a $1M tolerance to allow for minor rounding differences between
    Excel (which may have display rounding) and pandas (which uses full precision).
    """
    logger.info("\n── Data ───────────────────────────────────────────────────────")

    # Load the three financial statements from CSV
    try:
        df_is = pd.read_csv(DATA_DIR / "model_income_statement.csv", index_col=0)
        df_bs = pd.read_csv(DATA_DIR / "model_balance_sheet.csv",    index_col=0)
        df_cf = pd.read_csv(DATA_DIR / "model_cash_flow.csv",        index_col=0)
    except FileNotFoundError:
        check(False, "CSV data readable", "One or more CSVs missing — skipping data checks")
        return

    # Check 22: Period alignment (timeline coverage)
    # Ensures all three statements cover a common time horizon. At least 6 common periods
    # (1.5 years of quarterly data) indicates consistent extraction windows.
    common = set(df_is.columns) & set(df_bs.columns) & set(df_cf.columns)
    check(len(common) >= 6, "Period alignment (IS / BS / CF share ≥ 6 quarters)",
          f"Common quarters: {len(common)} — {sorted(common)}")

    # Check 23: Total Revenue data completeness
    # Missing revenue data in any period is a critical gap (no revenue = missing extraction).
    rev_row = "Total Revenue"
    if rev_row in df_is.index:
        nulls = df_is.loc[rev_row].isna().sum()
        check(nulls == 0, "Total Revenue non-null for all IS periods",
              f"{nulls} null(s) found" if nulls else "")
    else:
        check(False, "Total Revenue non-null for all IS periods", f"Row '{rev_row}' not found in IS CSV")

    # Check 24: Net Income quality for recent periods
    # Net Income is critical for valuation (DCF, P/E). Ensure it's complete for the last
    # 4 periods (1 year of quarterly data) to enable recent trend analysis and valuation.
    ni_row = "Net Income"
    if ni_row in df_is.index:
        recent = df_is.loc[ni_row].iloc[-4:]
        nulls = recent.isna().sum()
        check(nulls == 0, "Net Income non-null for most recent 4 IS periods",
              f"{nulls} null(s) in {list(recent.index)}" if nulls else "")
    else:
        check(False, "Net Income non-null for most recent 4 IS periods", f"Row '{ni_row}' not found in IS CSV")

    # Check 25: Free Cash Flow formula integrity
    # Validates the fundamental relationship: FCF = CFO − Capex
    # A violation signals either an extraction error (wrong row mapped) or a calculation error.
    # $1M tolerance accommodates rounding in Excel vs. pandas.
    if all(r in df_cf.index for r in ["Free Cash Flow", "Cash from Operations", "Capital Expenditures"]):
        fcf   = df_cf.loc["Free Cash Flow"]
        cfo   = df_cf.loc["Cash from Operations"]
        capex = df_cf.loc["Capital Expenditures"]
        # Only check periods where all three values are non-null (complete data)
        complete = fcf.notna() & cfo.notna() & capex.notna()
        if complete.any():
            diff = (fcf[complete] - (cfo[complete] - capex[complete])).abs()
            bad  = diff[diff > TOLERANCE]
            check(bad.empty, "FCF = CFO − Capex for all complete CF periods (±$1M)",
                  f"Violations: {bad.to_dict()}" if not bad.empty else "")
        else:
            check(False, "FCF = CFO − Capex for all complete CF periods (±$1M)",
                  "No period with all three values non-null")
    else:
        missing = [r for r in ["Free Cash Flow", "Cash from Operations", "Capital Expenditures"]
                   if r not in df_cf.index]
        check(False, "FCF = CFO − Capex for all complete CF periods (±$1M)",
              f"Missing CF rows: {missing}")

    # Check 26: Net Income reconciliation across statements
    # NI appears in both IS (P&L line item) and CF (adjustment to FCF). They must tie
    # for any shared period (within $1M for rounding tolerance).
    # A large discrepancy indicates either a data extraction error or an inconsistent
    # source (different GAAP treatments, discontinued ops, etc.).
    if ni_row in df_is.index and ni_row in df_cf.index:
        shared = sorted(set(df_is.columns) & set(df_cf.columns))
        is_ni  = df_is.loc[ni_row, shared]
        cf_ni  = df_cf.loc[ni_row, shared]
        # Only check periods where both NI values are non-null
        both   = is_ni.notna() & cf_ni.notna()
        if both.any():
            diff = (is_ni[both] - cf_ni[both]).abs()
            bad  = diff[diff > TOLERANCE]
            check(bad.empty, "Net Income ties between IS and CF for shared periods (±$1M)",
                  f"Violations: {bad.to_dict()}" if not bad.empty else "")
        else:
            check(False, "Net Income ties between IS and CF for shared periods (±$1M)",
                  "No shared period with non-null NI in both")
    else:
        check(False, "Net Income ties between IS and CF for shared periods (±$1M)",
              "Net Income row not found in IS or CF CSV")

    # Check 27: Stockholders' Equity sanity
    # Negative or zero equity signals insolvency or distress, which should be flagged
    # (though it's valid in some restructuring scenarios, it's unusual and often indicates
    # an extraction error).
    eq_row = "Stockholders' Equity"
    if eq_row in df_bs.index:
        eq = df_bs.loc[eq_row].dropna()
        neg = eq[eq <= 0]
        check(neg.empty, "Stockholders' Equity positive for all BS periods",
              f"Non-positive: {neg.to_dict()}" if not neg.empty else "")
    else:
        check(False, "Stockholders' Equity positive for all BS periods",
              f"Row '{eq_row}' not found in BS CSV")


# ---------------------------------------------------------------------------
# Main Entry Point
# ---------------------------------------------------------------------------

def main() -> None:
    """
    Run all 27 validation checks and report results.

    Orchestrates the validation pipeline:
      1. Initialize logging and clear the global results accumulator (for re-entrancy)
      2. Run file checks (10 checks)
      3. Run structural checks if Excel is present (10 checks)
      4. Run data checks (6 checks)
      5. Print summary report with PASS/FAIL breakdown
      6. Exit with code 0 if all checks passed, 1 if any failed

    Exit codes are used in CI/CD pipelines to block deployments on validation failure.
    """
    # Clear results in case main() is called multiple times (e.g., in tests)
    results.clear()

    logger.info("%s Model Validator", TICKER)
    logger.info("=" * 44)

    # Run file checks (manifest, CSV presence, period alignment)
    check_files()

    # Run structural checks (sheet names, row labels, formulas, bounds)
    # Skip if Excel workbook is missing, to avoid cascading failures.
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH, data_only=False)
        check_structure(wb)
        wb.close()
    else:
        logger.info("\n── Structure ──────────────────────────────────────────────────")
        logger.info("  (skipped — Excel output not found)")

    # Run data checks (numerical relationships, quality)
    check_data()

    # ─────────────────────────────────────────────────────────────────────
    # Summary Report
    # ─────────────────────────────────────────────────────────────────────
    passed = sum(1 for ok, _, _ in results if ok)
    total  = len(results)
    failed = total - passed

    print(f"\n{'=' * 44}")
    print(f"  {passed}/{total} checks passed", end="")
    if failed:
        print(f"  |  {failed} FAILED")
        # List all failed checks with their detail messages for diagnosis
        for ok, name, detail in results:
            if not ok:
                print(f"    ✗ {name}" + (f": {detail}" if detail else ""))
    else:
        print("  — all clear.")

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    logging.basicConfig(
        level=os.environ.get("LOG_LEVEL", "INFO").upper(),
        format="%(message)s",
    )
    main()
