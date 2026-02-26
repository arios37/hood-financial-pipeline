"""
Tests for the HOOD Financial Pipeline.

Test categories:
  - Integration: Full pipeline end-to-end (skip_extract=True)
  - Transform:   CSV data quality, row labels, period counts
  - Build:       Excel output structure, sheets, formulas, scenarios
  - Validate:    Validator catches known-good and known-bad states
  - Config:      Scenario YAML loading and bounds enforcement

Note:
    All tests rely on committed model CSVs in data/ so the SEC extraction
    step is not required.  The full pipeline (with extraction) requires
    network access and is exercised via ``make build`` or CI workflows.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

from config import (
    DATA_DIR,
    EXPECTED_SHEETS,
    OUTPUT_DIR,
    TICKER,
    IS_ROW,
    BS_ROW,
    CF_ROW,
)
from src.pipeline import run_pipeline
from src.hood_data_transform import main as transform
from src.build_excel_model import main as build_model, ASSUMPTIONS_SPEC

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

EXCEL_PATH = OUTPUT_DIR / f"{TICKER}_Financial_Model.xlsx"


@pytest.fixture(scope="session", autouse=True)
def run_pipeline_once():
    """Run the pipeline once for the entire test session.

    Executes transform → build → validate with skip_extract=True.
    All tests share the resulting Excel and CSV outputs.
    """
    run_pipeline(skip_extract=True)


@pytest.fixture()
def wb():
    """Load the generated Excel workbook (formulas, not computed values)."""
    workbook = load_workbook(EXCEL_PATH, data_only=False)
    yield workbook
    workbook.close()


@pytest.fixture()
def wb_values():
    """Load the generated Excel workbook with computed values (data_only=True)."""
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    yield workbook
    workbook.close()


@pytest.fixture()
def df_is():
    return pd.read_csv(DATA_DIR / "model_income_statement.csv", index_col=0)


@pytest.fixture()
def df_bs():
    return pd.read_csv(DATA_DIR / "model_balance_sheet.csv", index_col=0)


@pytest.fixture()
def df_cf():
    return pd.read_csv(DATA_DIR / "model_cash_flow.csv", index_col=0)


@pytest.fixture()
def manifest():
    with open(DATA_DIR / "manifest.json") as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------
# 1. Integration Tests
# ---------------------------------------------------------------------------


class TestIntegration:
    """End-to-end pipeline integration tests."""

    def test_pipeline_runs(self):
        """Pipeline executes transform → build → validate without exceptions.

        This is the original smoke test — kept for backward compatibility.
        The session-scoped fixture already ran the pipeline, so a second call
        just confirms re-entrancy.
        """
        run_pipeline(skip_extract=True)

    def test_excel_output_exists(self):
        """Pipeline produces the base-case Excel workbook."""
        assert EXCEL_PATH.exists(), f"Expected {EXCEL_PATH} to exist after pipeline run"

    def test_all_csvs_exist(self):
        """Pipeline produces all three model CSVs."""
        for name in ["model_income_statement.csv", "model_balance_sheet.csv", "model_cash_flow.csv"]:
            path = DATA_DIR / name
            assert path.exists(), f"Missing CSV: {path}"

    def test_manifest_exists(self):
        """Transform step produces manifest.json."""
        assert (DATA_DIR / "manifest.json").exists()


# ---------------------------------------------------------------------------
# 2. Transform / CSV Data Quality Tests
# ---------------------------------------------------------------------------


class TestTransformData:
    """Validate the model-ready CSVs produced by the transform step."""

    def test_is_required_rows(self, df_is):
        """Income Statement CSV contains all required line items."""
        required = [
            "Transaction-based Revenue",
            "Net Interest Revenue",
            "Other Revenue",
            "Total Revenue",
            "Operating Expenses",
            "Stock-Based Compensation",
            "Net Income",
        ]
        for row in required:
            assert row in df_is.index, f"IS missing row: {row}"

    def test_bs_required_rows(self, df_bs):
        """Balance Sheet CSV contains all required line items."""
        required = [
            "Cash & Cash Equivalents",
            "Restricted Cash",
            "Receivables",
            "Payables",
            "Total Debt",
            "Stockholders' Equity",
        ]
        for row in required:
            assert row in df_bs.index, f"BS missing row: {row}"

    def test_cf_required_rows(self, df_cf):
        """Cash Flow CSV contains all required line items."""
        required = [
            "Net Income",
            "Stock-Based Compensation",
            "Cash from Operations",
            "Capital Expenditures",
            "Free Cash Flow",
        ]
        for row in required:
            assert row in df_cf.index, f"CF missing row: {row}"

    def test_is_minimum_quarters(self, df_is):
        """IS has at least 8 quarters of data."""
        assert len(df_is.columns) >= 8, f"IS only has {len(df_is.columns)} quarters"

    def test_bs_minimum_quarters(self, df_bs):
        """BS has at least 6 quarters of data."""
        assert len(df_bs.columns) >= 6, f"BS only has {len(df_bs.columns)} quarters"

    def test_cf_minimum_quarters(self, df_cf):
        """CF has at least 6 quarters of data."""
        assert len(df_cf.columns) >= 6, f"CF only has {len(df_cf.columns)} quarters"

    def test_revenue_no_nulls(self, df_is):
        """Total Revenue has no null values across all periods."""
        nulls = df_is.loc["Total Revenue"].isna().sum()
        assert nulls == 0, f"Total Revenue has {nulls} null(s)"

    def test_revenue_all_positive(self, df_is):
        """Total Revenue is positive for every period."""
        rev = df_is.loc["Total Revenue"]
        non_positive = rev[rev <= 0]
        assert non_positive.empty, f"Non-positive revenue: {non_positive.to_dict()}"

    def test_equity_all_positive(self, df_bs):
        """Stockholders' Equity is positive for all periods."""
        eq = df_bs.loc["Stockholders' Equity"].dropna()
        neg = eq[eq <= 0]
        assert neg.empty, f"Non-positive equity: {neg.to_dict()}"

    def test_fcf_formula_holds(self, df_cf):
        """FCF = CFO - Capex for all complete periods (within $1M tolerance)."""
        fcf = df_cf.loc["Free Cash Flow"]
        cfo = df_cf.loc["Cash from Operations"]
        capex = df_cf.loc["Capital Expenditures"]
        complete = fcf.notna() & cfo.notna() & capex.notna()
        assert complete.any(), "No complete FCF/CFO/Capex period found"
        diff = (fcf[complete] - (cfo[complete] - capex[complete])).abs()
        bad = diff[diff > 1.0]
        assert bad.empty, f"FCF != CFO - Capex: {bad.to_dict()}"

    def test_net_income_ties_is_cf(self, df_is, df_cf):
        """Net Income matches between IS and CF for shared periods (within $1M)."""
        shared = sorted(set(df_is.columns) & set(df_cf.columns))
        assert len(shared) >= 6, f"Only {len(shared)} shared periods"
        is_ni = df_is.loc["Net Income", shared]
        cf_ni = df_cf.loc["Net Income", shared]
        both = is_ni.notna() & cf_ni.notna()
        diff = (is_ni[both] - cf_ni[both]).abs()
        bad = diff[diff > 1.0]
        assert bad.empty, f"NI mismatch IS vs CF: {bad.to_dict()}"

    def test_period_alignment(self, df_is, df_bs, df_cf):
        """IS, BS, and CF share at least 6 common quarters."""
        common = set(df_is.columns) & set(df_bs.columns) & set(df_cf.columns)
        assert len(common) >= 6, f"Only {len(common)} common quarters"

    def test_bs_periods_subset_of_is(self, df_is, df_bs):
        """All BS periods are also present in IS."""
        extra = set(df_bs.columns) - set(df_is.columns)
        assert not extra, f"BS has periods not in IS: {extra}"

    def test_cf_periods_subset_of_is(self, df_is, df_cf):
        """All CF periods are also present in IS."""
        extra = set(df_cf.columns) - set(df_is.columns)
        assert not extra, f"CF has periods not in IS: {extra}"


# ---------------------------------------------------------------------------
# 3. Manifest Tests
# ---------------------------------------------------------------------------


class TestManifest:
    """Validate manifest.json consistency with CSV outputs."""

    def test_manifest_has_all_statements(self, manifest):
        """Manifest contains entries for IS, BS, and CF."""
        stmts = manifest.get("statements", {})
        for key in ("IS", "BS", "CF"):
            assert key in stmts, f"Manifest missing statement key: {key}"

    def test_manifest_is_periods_match_csv(self, manifest, df_is):
        """Manifest IS periods match actual CSV columns."""
        expected = manifest["statements"]["IS"]["periods"]
        actual = list(df_is.columns)
        assert actual == expected, f"IS periods mismatch:\n  manifest: {expected}\n  CSV: {actual}"

    def test_manifest_bs_periods_match_csv(self, manifest, df_bs):
        expected = manifest["statements"]["BS"]["periods"]
        actual = list(df_bs.columns)
        assert actual == expected, f"BS periods mismatch"

    def test_manifest_cf_periods_match_csv(self, manifest, df_cf):
        expected = manifest["statements"]["CF"]["periods"]
        actual = list(df_cf.columns)
        assert actual == expected, f"CF periods mismatch"

    def test_manifest_has_timestamp(self, manifest):
        """Manifest includes a generated_at timestamp."""
        assert "generated_at" in manifest


# ---------------------------------------------------------------------------
# 4. Excel Structure Tests
# ---------------------------------------------------------------------------


class TestExcelStructure:
    """Validate the generated Excel workbook structure and formulas."""

    def test_all_seven_sheets_present(self, wb):
        """Workbook contains all 7 expected sheets."""
        for sheet_name in EXPECTED_SHEETS:
            assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"

    def test_no_extra_sheets(self, wb):
        """Workbook has no unexpected extra sheets."""
        extra = set(wb.sheetnames) - set(EXPECTED_SHEETS)
        assert not extra, f"Unexpected sheets: {extra}"

    def test_is_row_labels(self, wb):
        """Income Statement has correct row labels at expected positions."""
        ws = wb["Income Statement"]
        checks = {
            IS_ROW["Total Revenue"]: "Total Revenue",
            IS_ROW["Operating Income"]: "Operating Income",
            IS_ROW["Net Income"]: "Net Income",
        }
        for row, expected_label in checks.items():
            actual = ws.cell(row=row, column=1).value
            assert actual == expected_label, (
                f"IS row {row}: expected '{expected_label}', got '{actual}'"
            )

    def test_bs_row_labels(self, wb):
        """Balance Sheet has correct row labels at expected positions."""
        ws = wb["Balance Sheet"]
        checks = {
            BS_ROW["Cash"]: "Cash & Cash Equivalents",
            BS_ROW["Total Debt"]: "Total Debt",
            BS_ROW["Equity"]: "Stockholders' Equity",
        }
        for row, expected_label in checks.items():
            actual = ws.cell(row=row, column=1).value
            assert actual == expected_label, (
                f"BS row {row}: expected '{expected_label}', got '{actual}'"
            )

    def test_cf_row_labels(self, wb):
        """Cash Flow has correct row labels at expected positions."""
        ws = wb["Cash Flow"]
        checks = {
            CF_ROW["CFO"]: "Cash from Operations",
            CF_ROW["FCF"]: "Free Cash Flow",
        }
        for row, expected_label in checks.items():
            actual = ws.cell(row=row, column=1).value
            assert actual == expected_label, (
                f"CF row {row}: expected '{expected_label}', got '{actual}'"
            )

    def test_assumptions_sheet_has_inputs(self, wb):
        """Assumptions sheet has populated input cells."""
        ws = wb["Assumptions"]
        # Row 5, col B should be a growth rate value (first assumption)
        val = ws.cell(row=5, column=2).value
        assert val is not None, "Assumptions B5 is empty — expected a growth rate input"
        assert isinstance(val, (int, float)), f"Assumptions B5 should be numeric, got {type(val)}"


# ---------------------------------------------------------------------------
# 5. Excel Formula Tests
# ---------------------------------------------------------------------------


class TestExcelFormulas:
    """Verify forecast cells contain live Excel formulas, not hardcoded values."""

    @staticmethod
    def _find_fcst_col(ws, header_row=3):
        """Find the first forecast column by scanning for 'FY2026' header."""
        for cell in ws[header_row]:
            if str(cell.value or "").startswith("FY2026"):
                return cell.column
        pytest.fail(f"Could not find FY2026E header in row {header_row} of {ws.title}")

    def test_is_forecast_cells_are_formulas(self, wb):
        """IS forecast cells (revenue, OpEx, NI) are Excel formulas."""
        ws = wb["Income Statement"]
        fcst_col = self._find_fcst_col(ws)
        data_rows = [
            IS_ROW["Txn Revenue"],
            IS_ROW["Total Revenue"],
            IS_ROW["Operating Expenses"],
            IS_ROW["Net Income"],
        ]
        for row in data_rows:
            for offset in range(4):
                cell = ws.cell(row=row, column=fcst_col + offset)
                if cell.value is not None:
                    assert isinstance(cell.value, str) and cell.value.startswith("="), (
                        f"{cell.coordinate} is not a formula: {cell.value!r}"
                    )

    def test_bs_forecast_cells_are_formulas(self, wb):
        """BS forecast cells (Cash, Debt, Equity) are Excel formulas."""
        ws = wb["Balance Sheet"]
        fcst_col = self._find_fcst_col(ws)
        data_rows = [BS_ROW["Cash"], BS_ROW["Total Debt"], BS_ROW["Equity"]]
        for row in data_rows:
            for offset in range(4):
                cell = ws.cell(row=row, column=fcst_col + offset)
                if cell.value is not None:
                    assert isinstance(cell.value, str) and cell.value.startswith("="), (
                        f"{cell.coordinate} is not a formula: {cell.value!r}"
                    )

    def test_cf_forecast_cells_are_formulas(self, wb):
        """CF forecast cells (CFO, FCF) are Excel formulas."""
        ws = wb["Cash Flow"]
        fcst_col = self._find_fcst_col(ws)
        data_rows = [CF_ROW["CFO"], CF_ROW["FCF"]]
        for row in data_rows:
            for offset in range(4):
                cell = ws.cell(row=row, column=fcst_col + offset)
                if cell.value is not None:
                    assert isinstance(cell.value, str) and cell.value.startswith("="), (
                        f"{cell.coordinate} is not a formula: {cell.value!r}"
                    )

    def test_is_revenue_references_assumptions(self, wb):
        """IS FY2026E Txn Revenue formula references the Assumptions sheet."""
        ws = wb["Income Statement"]
        fcst_col = self._find_fcst_col(ws)
        formula = str(ws.cell(row=IS_ROW["Txn Revenue"], column=fcst_col).value or "")
        assert "Assumptions!" in formula, (
            f"FY2026E Txn Revenue does not reference Assumptions: {formula[:80]}"
        )

    def test_valuation_has_dcf_formulas(self, wb):
        """Valuation sheet contains formula cells in the DCF area."""
        ws = wb["Valuation"]
        val_cells = [ws.cell(row=r, column=2) for r in range(3, 8)]
        has_formula = any(
            isinstance(c.value, str) and c.value.startswith("=") for c in val_cells
        )
        assert has_formula, "No formula cells found in Valuation rows 3-7 col B"

    def test_sensitivity_cells_are_formulas(self, wb):
        """Sensitivity Analysis data cells (rows 7-11, cols B-F) are formulas."""
        ws = wb["Sensitivity Analysis"]
        for r in range(7, 12):
            for c in range(2, 7):
                cell = ws.cell(row=r, column=c)
                assert isinstance(cell.value, str) and cell.value.startswith("="), (
                    f"Sensitivity {cell.coordinate} is not a formula: {cell.value!r}"
                )


# ---------------------------------------------------------------------------
# 6. Assumption Bounds Tests
# ---------------------------------------------------------------------------


class TestAssumptionBounds:
    """Verify assumption values in the workbook are within valid ranges."""

    def test_growth_rates_in_bounds(self, wb):
        """Growth rate assumptions are between -50% and +200%."""
        ws = wb["Assumptions"]
        growth_rows = list(range(5, 9)) + list(range(11, 15)) + list(range(17, 21))
        for r in growth_rows:
            v = ws.cell(row=r, column=2).value
            if isinstance(v, (int, float)):
                assert -0.50 <= v <= 2.00, f"B{r}={v:.1%} outside growth bounds"

    def test_tax_rates_in_bounds(self, wb):
        """Tax rate assumptions are between 0% and 50%."""
        ws = wb["Assumptions"]
        for r in range(28, 32):
            v = ws.cell(row=r, column=2).value
            if isinstance(v, (int, float)):
                assert 0.0 <= v <= 0.50, f"B{r}={v:.1%} outside tax rate bounds"

    def test_wacc_in_bounds(self, wb):
        """WACC is between 5% and 30%."""
        ws = wb["Assumptions"]
        v = ws.cell(row=50, column=2).value
        if isinstance(v, (int, float)):
            assert 0.05 <= v <= 0.30, f"WACC={v:.1%} outside bounds"

    def test_terminal_growth_in_bounds(self, wb):
        """Terminal growth rate is between 0% and 10%."""
        ws = wb["Assumptions"]
        v = ws.cell(row=51, column=2).value
        if isinstance(v, (int, float)):
            assert 0.0 <= v <= 0.10, f"TGR={v:.1%} outside bounds"


# ---------------------------------------------------------------------------
# 7. Scenario Configuration Tests
# ---------------------------------------------------------------------------


class TestScenarios:
    """Validate scenario YAML loading and override mechanics."""

    @pytest.fixture()
    def all_scenarios(self):
        scenarios_path = Path(__file__).resolve().parents[1] / "config" / "scenarios.yaml"
        with open(scenarios_path) as fh:
            return yaml.safe_load(fh)

    def test_scenarios_yaml_loads(self, all_scenarios):
        """scenarios.yaml is valid YAML and parseable."""
        assert isinstance(all_scenarios, dict)

    def test_three_scenarios_defined(self, all_scenarios):
        """Bull, base, and bear scenarios are all defined."""
        for name in ("bull", "base", "bear"):
            assert name in all_scenarios, f"Missing scenario: {name}"

    def test_base_has_no_overrides(self, all_scenarios):
        """Base scenario is empty (all defaults)."""
        base = all_scenarios.get("base", {}) or {}
        assert len(base) == 0, f"Base scenario should be empty, got: {base}"

    def test_bull_overrides_are_numeric(self, all_scenarios):
        """All bull scenario override values are numeric."""
        bull = all_scenarios.get("bull", {}) or {}
        for k, v in bull.items():
            assert isinstance(v, (int, float)), f"Bull {k}={v!r} is not numeric"

    def test_bear_overrides_are_numeric(self, all_scenarios):
        """All bear scenario override values are numeric."""
        bear = all_scenarios.get("bear", {}) or {}
        for k, v in bear.items():
            assert isinstance(v, (int, float)), f"Bear {k}={v!r} is not numeric"

    def test_scenario_keys_are_valid_assumptions(self, all_scenarios):
        """All scenario override keys match real ASSUMPTIONS_SPEC keys."""
        valid_keys = {
            item[3]
            for item in ASSUMPTIONS_SPEC
            if item is not None and item[3] is not None
        }
        for scenario_name in ("bull", "bear"):
            overrides = all_scenarios.get(scenario_name, {}) or {}
            unknown = set(overrides) - valid_keys
            assert not unknown, (
                f"Scenario '{scenario_name}' has unknown keys: {sorted(unknown)}"
            )

    def test_bull_growth_higher_than_bear(self, all_scenarios):
        """Bull txn_growth_y1 is higher than bear txn_growth_y1."""
        bull = all_scenarios.get("bull", {}) or {}
        bear = all_scenarios.get("bear", {}) or {}
        if "txn_growth_y1" in bull and "txn_growth_y1" in bear:
            assert bull["txn_growth_y1"] > bear["txn_growth_y1"], (
                f"Bull growth ({bull['txn_growth_y1']}) should exceed bear ({bear['txn_growth_y1']})"
            )
