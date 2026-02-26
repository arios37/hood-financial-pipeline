"""
Generate PNG screenshots of key Excel model sheets for the README.

Uses openpyxl to read cell values/styles and Pillow to render a clean
spreadsheet-like image. Each sheet gets its own PNG in assets/.
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

EXCEL_PATH = Path("output/HOOD_Financial_Model.xlsx")
ASSETS_DIR = Path("assets")
ASSETS_DIR.mkdir(exist_ok=True)

# Sheets to screenshot and how many rows/cols to capture
SHEETS_CONFIG = {
    "Assumptions": {"max_rows": 45, "max_cols": 8, "filename": "assumptions_sheet.png"},
    "Income Statement": {"max_rows": 30, "max_cols": 20, "filename": "income_statement.png"},
    "Sensitivity Analysis": {"max_rows": 35, "max_cols": 12, "filename": "sensitivity_analysis.png"},
    "Valuation": {"max_rows": 35, "max_cols": 10, "filename": "valuation_analysis.png"},
}

# Rendering constants
ROW_HEIGHT = 22
HEADER_HEIGHT = 28
PADDING_X = 8
PADDING_Y = 4
MIN_COL_WIDTH = 60
MAX_COL_WIDTH = 180


def hex_to_rgb(hex_color: str) -> tuple:
    """Convert hex color string to RGB tuple."""
    if not hex_color or hex_color == "00000000" or len(hex_color) < 6:
        return (255, 255, 255)
    # Take last 6 chars (skip alpha if present)
    h = hex_color[-6:]
    try:
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    except ValueError:
        return (255, 255, 255)


def get_font(bold=False, size=11):
    """Get a font, falling back to default if needed."""
    try:
        if bold:
            return ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", size)
        return ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", size)
    except (OSError, IOError):
        return ImageFont.load_default()


def format_value(value, number_format=None):
    """Format cell value for display."""
    if value is None:
        return ""
    if isinstance(value, str):
        # Truncate formulas for display
        if value.startswith("="):
            return "=formula"
        return value
    if isinstance(value, (int, float)):
        if number_format and "%" in str(number_format):
            return f"{value:.1%}"
        if number_format and "$" in str(number_format):
            if abs(value) >= 1_000_000:
                return f"${value/1_000_000:,.0f}M"
            elif abs(value) >= 1_000:
                return f"${value:,.0f}"
            else:
                return f"${value:.2f}"
        if number_format and "x" in str(number_format).lower():
            return f"{value:.1f}x"
        if isinstance(value, float):
            if abs(value) < 0.01 and value != 0:
                return f"{value:.4f}"
            if abs(value) >= 1_000_000:
                return f"{value/1_000_000:,.1f}M"
            if abs(value) >= 1_000:
                return f"{value:,.0f}"
            return f"{value:.2f}"
        if abs(value) >= 1_000:
            return f"{value:,}"
        return str(value)
    return str(value)


def render_sheet(ws, ws_values, config: dict, filename: str):
    """Render a worksheet to a PNG image."""
    max_rows = config["max_rows"]
    max_cols = config["max_cols"]

    font_regular = get_font(bold=False, size=11)
    font_bold = get_font(bold=True, size=11)
    font_header = get_font(bold=True, size=12)

    # First pass: determine column widths from content
    col_widths = []
    for col_idx in range(1, max_cols + 1):
        max_w = MIN_COL_WIDTH
        for row_idx in range(1, min(max_rows + 1, ws_values.max_row + 1)):
            cell = ws_values.cell(row=row_idx, column=col_idx)
            style_cell = ws.cell(row=row_idx, column=col_idx)
            text = format_value(cell.value, style_cell.number_format)
            if text:
                f = font_bold if (style_cell.font and style_cell.font.bold) else font_regular
                bbox = f.getbbox(text)
                tw = bbox[2] - bbox[0] if bbox else 0
                max_w = max(max_w, tw + PADDING_X * 2)
        col_widths.append(min(max_w, MAX_COL_WIDTH))

    # Calculate image dimensions
    total_width = sum(col_widths) + 2  # +2 for border
    actual_rows = min(max_rows, ws_values.max_row)
    total_height = actual_rows * ROW_HEIGHT + 2

    # Create image
    img = Image.new("RGB", (total_width, total_height), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    # Render cells
    y = 0
    for row_idx in range(1, actual_rows + 1):
        x = 1
        row_h = ROW_HEIGHT

        for col_idx in range(1, max_cols + 1):
            cw = col_widths[col_idx - 1]
            cell = ws_values.cell(row=row_idx, column=col_idx)
            style_cell = ws.cell(row=row_idx, column=col_idx)

            # Background fill
            bg = (255, 255, 255)
            if style_cell.fill and style_cell.fill.fgColor and style_cell.fill.fgColor.rgb:
                rgb = style_cell.fill.fgColor.rgb
                if isinstance(rgb, str) and len(rgb) >= 6:
                    bg = hex_to_rgb(rgb)

            draw.rectangle([x, y, x + cw - 1, y + row_h - 1], fill=bg)

            # Cell text
            text = format_value(cell.value, style_cell.number_format)
            if text:
                # Font color
                fc = (0, 0, 0)
                if style_cell.font and style_cell.font.color and style_cell.font.color.rgb:
                    rgb = style_cell.font.color.rgb
                    if isinstance(rgb, str) and len(rgb) >= 6:
                        fc = hex_to_rgb(rgb)

                is_bold = style_cell.font and style_cell.font.bold
                f = font_bold if is_bold else font_regular

                # Truncate text to fit
                bbox = f.getbbox(text)
                tw = bbox[2] - bbox[0] if bbox else 0
                while tw > cw - PADDING_X * 2 and len(text) > 3:
                    text = text[:-4] + "..."
                    bbox = f.getbbox(text)
                    tw = bbox[2] - bbox[0] if bbox else 0

                # Right-align numbers, left-align text
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    tx = x + cw - PADDING_X - tw
                else:
                    tx = x + PADDING_X

                ty = y + PADDING_Y
                draw.text((tx, ty), text, fill=fc, font=f)

            # Cell border (light gray)
            draw.rectangle([x, y, x + cw - 1, y + row_h - 1], outline=(220, 220, 220))

            x += cw
        y += row_h

    # Save
    output_path = ASSETS_DIR / filename
    img.save(output_path, "PNG", optimize=True)
    print(f"  Saved: {output_path} ({img.width}x{img.height})")
    return output_path


def main():
    print("Loading workbook (formulas)...")
    wb = load_workbook(EXCEL_PATH, data_only=False)
    print("Loading workbook (values)...")
    wb_values = load_workbook(EXCEL_PATH, data_only=True)

    for sheet_name, config in SHEETS_CONFIG.items():
        print(f"\nRendering: {sheet_name}")
        if sheet_name in wb.sheetnames:
            render_sheet(
                wb[sheet_name],
                wb_values[sheet_name],
                config,
                config["filename"],
            )
        else:
            print(f"  WARNING: Sheet '{sheet_name}' not found!")

    wb.close()
    wb_values.close()
    print("\nDone! Screenshots saved to assets/")


if __name__ == "__main__":
    main()
